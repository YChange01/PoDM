from __future__ import annotations

import hashlib
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from promote_reviewed_baseline import promote_reviewed_baseline  # noqa: E402


def write_reviewed_workbook(path: Path, note: str = "人工确认") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "共有接口"
    ws.append(
        [
            "confidence",
            "bmc_section",
            "bmc_title",
            "bmc_method",
            "bmc_uri",
            "podm_section",
            "podm_title",
            "podm_method",
            "podm_uri",
            "复核备注",
        ]
    )
    ws.append(
        [
            "high",
            "3.1.1",
            "BMC查询",
            "GET",
            "/redfish/v1",
            "4.1.1",
            "PoDM查询",
            "GET",
            "/redfish/v1",
            note,
        ]
    )
    ws = wb.create_sheet("BMC独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    ws.append(["3.2.1", "BMC独有", "GET", "/redfish/v1/BmcOnly", ""])
    ws = wb.create_sheet("PoDM独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    wb.save(path)


class PromoteReviewedBaselineTest(unittest.TestCase):
    def test_promotes_workbook_and_writes_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "new_summary.xlsx"
            baseline_dir = root / "baseline"

            write_reviewed_workbook(source)

            manifest = promote_reviewed_baseline(source, baseline_dir, baseline_date="20260610")
            target = baseline_dir / "reviewed_baseline.xlsx"
            manifest_path = baseline_dir / "reviewed_baseline_manifest.json"

            self.assertTrue(target.is_file())
            self.assertTrue(manifest_path.is_file())
            self.assertEqual(hashlib.sha256(target.read_bytes()).hexdigest(), manifest["sha256"])

            data = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual("20260610", data["baseline_date"])
            self.assertEqual(1, data["sheets"]["共有接口"]["rows"])
            self.assertEqual({"复核备注": 1}, data["sheets"]["共有接口"]["note_counts"])
            self.assertEqual(1, data["sheets"]["BMC独有"]["rows"])
            self.assertEqual(0, data["sheets"]["PoDM独有"]["rows"])

    def test_archives_existing_baseline_before_replacing_it(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "new_summary.xlsx"
            baseline_dir = root / "baseline"
            baseline_dir.mkdir()
            old_workbook = baseline_dir / "reviewed_baseline.xlsx"
            old_manifest = baseline_dir / "reviewed_baseline_manifest.json"

            write_reviewed_workbook(old_workbook, note="旧基线")
            old_manifest.write_text(
                json.dumps({"sha256": hashlib.sha256(old_workbook.read_bytes()).hexdigest()}),
                encoding="utf-8",
            )
            old_manifest_text = old_manifest.read_text(encoding="utf-8")
            write_reviewed_workbook(source, note="新基线")

            manifest = promote_reviewed_baseline(
                source,
                baseline_dir,
                baseline_date="20260610",
                backup_timestamp="20260610T010203Z",
            )

            backup_dir = baseline_dir / "backup_20260610T010203Z"
            self.assertEqual(str(backup_dir), manifest["previous_baseline_backup"])
            self.assertTrue((backup_dir / "reviewed_baseline.xlsx").is_file())
            self.assertTrue((backup_dir / "reviewed_baseline_manifest.json").is_file())
            self.assertEqual(
                old_manifest_text,
                (backup_dir / "reviewed_baseline_manifest.json").read_text(encoding="utf-8"),
            )
            self.assertEqual(hashlib.sha256(old_workbook.read_bytes()).hexdigest(), manifest["sha256"])

    def test_rejects_workbook_without_required_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "bad.xlsx"
            baseline_dir = root / "baseline"

            wb = Workbook()
            ws = wb.active
            ws.title = "共有接口"
            ws.append(["confidence", "bmc_section", "复核备注"])
            wb.save(source)

            with self.assertRaises(SystemExit):
                promote_reviewed_baseline(source, baseline_dir, baseline_date="20260610")

            self.assertFalse((baseline_dir / "reviewed_baseline.xlsx").exists())


if __name__ == "__main__":
    unittest.main()

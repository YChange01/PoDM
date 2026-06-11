from __future__ import annotations

import hashlib
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from promote_podm_bmc_baseline import promote_podm_bmc_baseline  # noqa: E402


def write_podm_bmc_workbook(path: Path, note: str = "人工确认") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "共有接口"
    ws.append(
        [
            "confidence",
            "bmc_section",
            "bmc_title",
            "bmc_method",
            "bmc_uri",
            "podm_section",
            "podm_title",
            "podm_method",
            "podm_uri",
            "复核备注",
        ]
    )
    ws.append(
        [
            "high",
            "3.1.1",
            "BMC查询",
            "GET",
            "/redfish/v1",
            "4.1.1",
            "PoDM查询",
            "GET",
            "/redfish/v1",
            note,
        ]
    )
    ws = wb.create_sheet("BMC独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    ws.append(["3.2.1", "BMC独有", "GET", "/redfish/v1/BmcOnly", ""])
    ws = wb.create_sheet("PoDM独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    wb.save(path)


class PromotePodmBmcBaselineTest(unittest.TestCase):
    def test_promotes_workbook_and_writes_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "podm_bmc_summary.xlsx"
            baseline_dir = root / "baseline"

            write_podm_bmc_workbook(source)

            manifest = promote_podm_bmc_baseline(source, baseline_dir, baseline_date="20260610")
            target = baseline_dir / "podm_bmc" / "baseline.xlsx"
            manifest_path = baseline_dir / "podm_bmc" / "manifest.json"

            self.assertTrue(target.is_file())
            self.assertTrue(manifest_path.is_file())
            self.assertEqual(hashlib.sha256(target.read_bytes()).hexdigest(), manifest["sha256"])

            data = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual("20260610", data["baseline_date"])
            self.assertEqual(1, data["sheets"]["共有接口"]["rows"])
            self.assertEqual({"复核备注": 1}, data["sheets"]["共有接口"]["note_counts"])
            self.assertEqual(1, data["sheets"]["BMC独有"]["rows"])
            self.assertEqual(0, data["sheets"]["PoDM独有"]["rows"])

    def test_archives_existing_baseline_before_replacing_it(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "podm_bmc_summary.xlsx"
            baseline_dir = root / "baseline"
            workflow_dir = baseline_dir / "podm_bmc"
            workflow_dir.mkdir(parents=True)
            old_workbook = workflow_dir / "baseline.xlsx"
            old_manifest = workflow_dir / "manifest.json"

            write_podm_bmc_workbook(old_workbook, note="旧基线")
            old_sha256 = hashlib.sha256(old_workbook.read_bytes()).hexdigest()
            old_manifest.write_text(
                json.dumps({"workbook": str(old_workbook), "sha256": old_sha256}),
                encoding="utf-8",
            )
            write_podm_bmc_workbook(source, note="新基线")

            manifest = promote_podm_bmc_baseline(
                source,
                baseline_dir,
                baseline_date="20260610",
                backup_timestamp="20260610T010203Z",
            )

            backup_dir = workflow_dir / "backups" / "20260610T010203Z"
            self.assertEqual(str(backup_dir), manifest["previous_baseline_backup"])
            backup_workbook = backup_dir / "baseline_20260610T010203Z.xlsx"
            backup_manifest_path = backup_dir / "manifest_20260610T010203Z.json"
            self.assertTrue(backup_workbook.is_file())
            self.assertTrue(backup_manifest_path.is_file())
            backup_manifest = json.loads(backup_manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(str(backup_workbook), backup_manifest["workbook"])
            self.assertEqual(hashlib.sha256(backup_workbook.read_bytes()).hexdigest(), backup_manifest["sha256"])
            self.assertEqual(hashlib.sha256(old_workbook.read_bytes()).hexdigest(), manifest["sha256"])

    def test_rejects_workbook_without_required_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "bad.xlsx"
            baseline_dir = root / "baseline"

            wb = Workbook()
            ws = wb.active
            ws.title = "共有接口"
            ws.append(["confidence", "bmc_section", "复核备注"])
            wb.save(source)

            with self.assertRaises(SystemExit):
                promote_podm_bmc_baseline(source, baseline_dir, baseline_date="20260610")

            self.assertFalse((baseline_dir / "podm_bmc" / "baseline.xlsx").exists())


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import hashlib
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from compare_podm_bmc_to_baseline import compare_podm_bmc_to_baseline  # noqa: E402
from compare_podm_bmc_to_baseline import main as update_main  # noqa: E402


def write_yaml(path: Path, items: list[dict[str, str]]) -> None:
    lines = ["interfaces:"]
    for index, item in enumerate(items, start=1):
        lines.extend(
            [
                f"- index: {index}",
                f"  section: {item['section']}",
                f"  title: {item['title']}",
                f"  method: {item['method']}",
                f"  uri: {item['uri']}",
            ]
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_baseline(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "共有接口"
    ws.append(
        [
            "confidence",
            "bmc_section",
            "bmc_title",
            "bmc_method",
            "bmc_uri",
            "podm_section",
            "podm_title",
            "podm_method",
            "podm_uri",
            "复核备注",
        ]
    )
    ws.append(
        [
            "medium",
            "3.6.15",
            "SSH公钥删除",
            "POST",
            "https://device_ip/redfish/v1/AccountService/Accounts/account_id/Oem/Huawei/Public/Actions/Account.DeleteSSHPublicKey",
            "4.10.3.1.10",
            "SSH公钥删除",
            "DELETE",
            "/redfish/v1/AccountService/Accounts/account_id/Oem/{OemVendor}/Public/Actions/Account.DeleteSSHPublicKey",
            "",
        ]
    )
    ws.append(
        [
            "high",
            "3.9.9",
            "修改事件订阅资源",
            "PATCH",
            "https://device_ip/redfish/v1/EventService/Subscriptions/id",
            "4.7.9",
            "修改事件订阅资源",
            "PATCH",
            "/redfish/v1/EventService/Subscriptions/{id}",
            "",
        ]
    )
    ws = wb.create_sheet("BMC独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    ws.append(
        [
            "3.2.70",
            "校验VNC密码",
            "POST",
            "https://device_ip/redfish/v1/Managers/manager_id/VncService/Actions/VncService.VerifyPassword",
            "",
        ]
    )
    ws = wb.create_sheet("PoDM独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    ws.append(["4.2.8", "备份刀片配置文件", "POST", "/redfish/v1/Managers/{manager_id}/Actions/Oem/HuaweiPublic/Manager.BackupConfiguration", ""])
    wb.save(path)


class UpdateInterfaceSummaryFromBaselineTest(unittest.TestCase):
    def test_preserves_baseline_pairs_and_marks_additions_and_deletions(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline.xlsx"
            bmc_yaml = root / "bmc.interface-list.yaml"
            podm_yaml = root / "podm.interface-list.yaml"
            output = root / "updated.xlsx"

            write_baseline(baseline)
            write_yaml(
                bmc_yaml,
                [
                    {
                        "section": "3.6.15",
                        "title": "SSH公钥删除",
                        "method": "POST",
                        "uri": "https://device_ip/redfish/v1/AccountService/Accounts/account_id/Oem/Huawei/Public/Actions/Account.DeleteSSHPublicKey",
                    },
                    {
                        "section": "3.9.9",
                        "title": "修改事件订阅资源",
                        "method": "PATCH",
                        "uri": "https://device_ip/redfish/v1/EventService/Subscriptions/subscription_id",
                    },
                    {
                        "section": "3.20.1",
                        "title": "查询新资源",
                        "method": "GET",
                        "uri": "https://device_ip/redfish/v1/NewResource",
                    },
                    {
                        "section": "3.20.2",
                        "title": "BMC新增独有",
                        "method": "POST",
                        "uri": "https://device_ip/redfish/v1/Managers/manager_id/Actions/Oem/Huawei/Public/OnlyBmc.Action",
                    },
                ],
            )
            write_yaml(
                podm_yaml,
                [
                    {
                        "section": "4.10.3.1.10",
                        "title": "SSH公钥删除",
                        "method": "DELETE",
                        "uri": "/redfish/v1/AccountService/Accounts/account_id/Oem/{OemVendor}/Public/Actions/Account.DeleteSSHPublicKey",
                    },
                    {
                        "section": "4.7.9",
                        "title": "修改事件订阅资源",
                        "method": "PATCH",
                        "uri": "/redfish/v1/EventService/Subscriptions/{subscription_id}",
                    },
                    {
                        "section": "4.20.1",
                        "title": "查询新资源",
                        "method": "GET",
                        "uri": "/redfish/v1/NewResource",
                    },
                ],
            )

            compare_podm_bmc_to_baseline(baseline, bmc_yaml, podm_yaml, output)

            wb = load_workbook(output, read_only=True, data_only=True)
            common_rows = list(wb["共有接口"].iter_rows(min_row=2, values_only=True))
            bmc_only_rows = list(wb["BMC独有"].iter_rows(min_row=2, values_only=True))
            podm_only_rows = list(wb["PoDM独有"].iter_rows(min_row=2, values_only=True))

            self.assertEqual(3, len(common_rows))
            self.assertIn(("3.6.15", "SSH公钥删除", "POST", "4.10.3.1.10", "SSH公钥删除", "DELETE"), {
                (row[1], row[2], row[3], row[5], row[6], row[7]) for row in common_rows
            })
            new_common = next(row for row in common_rows if row[1] == "3.20.1")
            self.assertIn("新增", new_common[9])
            changed_common = next(row for row in common_rows if row[1] == "3.9.9")
            self.assertIn("变更", changed_common[9])
            self.assertIn("subscription_id", changed_common[9])

            self.assertEqual(2, len(bmc_only_rows))
            deleted_bmc = next(row for row in bmc_only_rows if row[0] == "3.2.70")
            self.assertIn("删除", deleted_bmc[4])
            added_bmc = next(row for row in bmc_only_rows if row[0] == "3.20.2")
            self.assertIn("新增", added_bmc[4])

            self.assertEqual(1, len(podm_only_rows))
            deleted_podm = podm_only_rows[0]
            self.assertEqual("4.2.8", deleted_podm[0])
            self.assertIn("删除", deleted_podm[4])

    def test_cli_writes_report_and_validates_baseline_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "baseline.xlsx"
            manifest = root / "baseline_manifest.json"
            bmc_yaml = root / "bmc.interface-list.yaml"
            podm_yaml = root / "podm.interface-list.yaml"
            output = root / "updated.xlsx"
            report = root / "podm_bmc_update_report.json"

            write_baseline(baseline)
            manifest.write_text(
                json.dumps(
                    {
                        "workbook": str(baseline),
                        "sha256": hashlib.sha256(baseline.read_bytes()).hexdigest(),
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            write_yaml(
                bmc_yaml,
                [
                    {
                        "section": "3.6.15",
                        "title": "SSH公钥删除",
                        "method": "POST",
                        "uri": "https://device_ip/redfish/v1/AccountService/Accounts/account_id/Oem/Huawei/Public/Actions/Account.DeleteSSHPublicKey",
                    },
                    {
                        "section": "3.9.9",
                        "title": "修改事件订阅资源",
                        "method": "PATCH",
                        "uri": "https://device_ip/redfish/v1/EventService/Subscriptions/subscription_id",
                    },
                    {
                        "section": "3.20.1",
                        "title": "查询新资源",
                        "method": "GET",
                        "uri": "https://device_ip/redfish/v1/NewResource",
                    },
                    {
                        "section": "3.20.2",
                        "title": "BMC新增独有",
                        "method": "POST",
                        "uri": "https://device_ip/redfish/v1/Managers/manager_id/Actions/Oem/Huawei/Public/OnlyBmc.Action",
                    },
                ],
            )
            write_yaml(
                podm_yaml,
                [
                    {
                        "section": "4.10.3.1.10",
                        "title": "SSH公钥删除",
                        "method": "DELETE",
                        "uri": "/redfish/v1/AccountService/Accounts/account_id/Oem/{OemVendor}/Public/Actions/Account.DeleteSSHPublicKey",
                    },
                    {
                        "section": "4.7.9",
                        "title": "修改事件订阅资源",
                        "method": "PATCH",
                        "uri": "/redfish/v1/EventService/Subscriptions/{subscription_id}",
                    },
                    {
                        "section": "4.20.1",
                        "title": "查询新资源",
                        "method": "GET",
                        "uri": "/redfish/v1/NewResource",
                    },
                ],
            )

            update_main(
                [
                    "--baseline",
                    str(baseline),
                    "--baseline-manifest",
                    str(manifest),
                    "--bmc-yaml",
                    str(bmc_yaml),
                    "--podm-yaml",
                    str(podm_yaml),
                    "--output",
                    str(output),
                    "--report",
                    str(report),
                ]
            )

            self.assertTrue(output.is_file())
            data = json.loads(report.read_text(encoding="utf-8"))
            self.assertEqual(1, data["summary"]["common"]["added"])
            self.assertEqual(1, data["summary"]["common"]["changed"])
            self.assertEqual(1, data["summary"]["bmc_only"]["added"])
            self.assertEqual(1, data["summary"]["bmc_only"]["deleted"])
            self.assertEqual(1, data["summary"]["podm_only"]["deleted"])
            self.assertEqual(1, data["summary"]["review_needed"]["bmc_only"])

            manifest.write_text(
                json.dumps({"workbook": str(baseline), "sha256": "bad"}, ensure_ascii=False),
                encoding="utf-8",
            )
            with self.assertRaises(SystemExit):
                update_main(
                    [
                        "--baseline",
                        str(baseline),
                        "--baseline-manifest",
                        str(manifest),
                        "--bmc-yaml",
                        str(bmc_yaml),
                        "--podm-yaml",
                        str(podm_yaml),
                        "--output",
                        str(output),
                    ]
                )


if __name__ == "__main__":
    unittest.main()

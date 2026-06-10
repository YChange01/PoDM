from __future__ import annotations

import hashlib
import json
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import extract_resource_tree_interfaces as resource_tree_extractor  # noqa: E402
from extract_resource_tree_interfaces import extract as extract_resource_tree  # noqa: E402
from extract_resource_tree_interfaces import main as extract_resource_tree_main  # noqa: E402
from _defaults import PODM_DOCX_NAME  # noqa: E402
from promote_resource_tree_baseline import promote_resource_tree_baseline  # noqa: E402
from update_resource_tree_summary_from_baseline import (  # noqa: E402
    update_resource_tree_summary,
)


def write_yaml(path: Path, items: list[dict[str, str]]) -> None:
    lines = ["interfaces:"]
    for index, item in enumerate(items, start=1):
        lines.extend(
            [
                f"- index: {index}",
                f"  section: {item['section']}",
                f"  title: {item['title']}",
                f"  method: {item['method']}",
                f"  uri: {item['uri']}",
            ]
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_docx_lines(path: Path, lines: list[str]) -> None:
    ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    paragraphs = [
        f'<w:p><w:r><w:t xml:space="preserve">{line}</w:t></w:r></w:p>'
        for line in lines
    ]
    document = (
        f'<w:document xmlns:w="{ns}"><w:body>'
        + "".join(paragraphs)
        + "</w:body></w:document>"
    )
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("word/document.xml", document)


def write_resource_tree_baseline(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "共有接口"
    ws.append(
        [
            "confidence",
            "resource_section",
            "resource_title",
            "resource_method",
            "resource_uri",
            "podm_section",
            "podm_title",
            "podm_method",
            "podm_uri",
            "复核备注",
        ]
    )
    ws.append(
        [
            "high",
            "Managers资源",
            "/redfish/v1/Managers/manager_id",
            "GET",
            "/redfish/v1/Managers/manager_id",
            "4.2.1",
            "查询管理资源",
            "GET",
            "/redfish/v1/Managers/{manager_id}",
            "",
        ]
    )
    ws = wb.create_sheet("资源树独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    ws.append(["Managers资源", "/redfish/v1/Managers/old", "GET", "/redfish/v1/Managers/old", ""])
    ws = wb.create_sheet("PoDM独有")
    ws.append(["section", "title", "method", "uri", "差异备注"])
    ws.append(["4.9.9", "旧PoDM独有", "POST", "/redfish/v1/OldPoDM", ""])
    wb.save(path)


class ResourceTreeWorkflowTest(unittest.TestCase):
    def test_extract_resource_tree_splits_multi_method_rows(self) -> None:
        text = "\n".join(
            [
                "3 Redfish接口资源树",
                "表3-1 Redfish资源树",
                "URL\t允许操作",
                "公共固定资源",
                "/redfish/v1\tGET/PATCH",
                "/redfish/v1/JSONSchemas\t",
                "Managers资源",
                "redfish/v1/Managers/manager_id/SmsService\tGET/PATCH",
                "/redfish/v1/Systems/system_id/Storages/storage_id/Volumes/volume_id\tGET/PATCH/Delete",
                "/redfish/v1/EventService/Subscriptions/subscription_id\tGET/DELETE",
                "4 下一个章节",
            ]
        )

        items = extract_resource_tree(text)

        self.assertEqual(
            [
                ("GET", "/redfish/v1"),
                ("PATCH", "/redfish/v1"),
                ("GET", "/redfish/v1/Managers/manager_id/SmsService"),
                ("PATCH", "/redfish/v1/Managers/manager_id/SmsService"),
                ("GET", "/redfish/v1/Systems/system_id/Storages/storage_id/Volumes/volume_id"),
                ("PATCH", "/redfish/v1/Systems/system_id/Storages/storage_id/Volumes/volume_id"),
                ("DELETE", "/redfish/v1/Systems/system_id/Storages/storage_id/Volumes/volume_id"),
                ("GET", "/redfish/v1/EventService/Subscriptions/subscription_id"),
                ("DELETE", "/redfish/v1/EventService/Subscriptions/subscription_id"),
            ],
            [(item.method, item.uri) for item in items],
        )
        self.assertEqual("公共固定资源", items[0].section)
        self.assertEqual("Managers资源", items[2].section)

    def test_extract_resource_tree_cli_writes_yaml(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "resource_tree.txt"
            output = root / "resource_tree.yaml"
            source.write_text(
                "3 Redfish接口资源树\nURL\t允许操作\nManagers资源\n/redfish/v1/Managers\tGET/POST\n",
                encoding="utf-8",
            )

            extract_resource_tree_main([str(source), str(output)])

            content = output.read_text(encoding="utf-8")
            self.assertIn("method: GET", content)
            self.assertIn("method: POST", content)
            self.assertIn("uri: /redfish/v1/Managers", content)

    def test_extract_resource_tree_cli_accepts_date_argument(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            data_dir = root / "data" / "20260609"
            data_dir.mkdir(parents=True)
            docx = data_dir / PODM_DOCX_NAME
            write_docx_lines(
                docx,
                [
                    "3 Redfish接口资源树",
                    "URL\t允许操作",
                    "Managers资源",
                    "/redfish/v1/Managers\tGET/POST",
                ],
            )

            original_data_dir = resource_tree_extractor.DATA_DIR
            original_output_dir = resource_tree_extractor.OUTPUT_DIR
            try:
                resource_tree_extractor.DATA_DIR = root / "data"
                resource_tree_extractor.OUTPUT_DIR = root / "output"
                extract_resource_tree_main(["20260609"])
            finally:
                resource_tree_extractor.DATA_DIR = original_data_dir
                resource_tree_extractor.OUTPUT_DIR = original_output_dir

            output = root / "output" / "20260609" / f"{Path(PODM_DOCX_NAME).stem}.resource-tree.interface-list.yaml"
            content = output.read_text(encoding="utf-8")
            self.assertIn("method: GET", content)
            self.assertIn("method: POST", content)

    def test_compare_resource_tree_with_podm_and_baseline(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            baseline = root / "resource_tree_baseline.xlsx"
            resource_yaml = root / "resource_tree.yaml"
            podm_yaml = root / "podm.yaml"
            output = root / "resource_tree_summary.xlsx"

            write_resource_tree_baseline(baseline)
            write_yaml(
                resource_yaml,
                [
                    {
                        "section": "Managers资源",
                        "title": "/redfish/v1/Managers/manager_id",
                        "method": "GET",
                        "uri": "/redfish/v1/Managers/manager_id",
                    },
                    {
                        "section": "Managers资源",
                        "title": "/redfish/v1/Managers/manager_id",
                        "method": "PATCH",
                        "uri": "/redfish/v1/Managers/manager_id",
                    },
                    {
                        "section": "Managers资源",
                        "title": "/redfish/v1/Managers/manager_id/Actions/Manager.Reset",
                        "method": "POST",
                        "uri": "/redfish/v1/Managers/manager_id/Actions/Manager.Reset",
                    },
                ],
            )
            write_yaml(
                podm_yaml,
                [
                    {
                        "section": "4.2.1",
                        "title": "查询管理资源",
                        "method": "GET",
                        "uri": "/redfish/v1/Managers/{manager_id}",
                    },
                    {
                        "section": "4.2.2",
                        "title": "修改管理资源",
                        "method": "PATCH",
                        "uri": "/redfish/v1/Managers/{manager_id}",
                    },
                    {
                        "section": "4.20.1",
                        "title": "PoDM新增独有",
                        "method": "DELETE",
                        "uri": "/redfish/v1/PoDMOnly",
                    },
                ],
            )

            report = update_resource_tree_summary(baseline, resource_yaml, podm_yaml, output)

            self.assertEqual(2, report["summary"]["common"]["rows"])
            self.assertEqual(1, report["summary"]["common"]["added"])
            self.assertEqual(2, report["summary"]["resource_tree_only"]["rows"])
            self.assertEqual(1, report["summary"]["resource_tree_only"]["deleted"])
            self.assertEqual(2, report["summary"]["podm_only"]["rows"])
            self.assertEqual(1, report["summary"]["podm_only"]["deleted"])

            wb = load_workbook(output, read_only=True, data_only=True)
            self.assertEqual(
                [
                    "confidence",
                    "resource_section",
                    "resource_title",
                    "resource_method",
                    "resource_uri",
                    "podm_section",
                    "podm_title",
                    "podm_method",
                    "podm_uri",
                    "复核备注",
                ],
                [cell.value for cell in next(wb["共有接口"].iter_rows(min_row=1, max_row=1))],
            )
            self.assertIn("资源树独有", wb.sheetnames)
            common_notes = [row[9] for row in wb["共有接口"].iter_rows(min_row=2, values_only=True)]
            self.assertTrue(any(note and "新增" in note for note in common_notes))
            resource_only_notes = [row[4] for row in wb["资源树独有"].iter_rows(min_row=2, values_only=True)]
            self.assertTrue(any(note and "删除" in note for note in resource_only_notes))
            wb.close()

    def test_promotes_resource_tree_baseline_and_archives_existing_one(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "resource_tree_summary.xlsx"
            baseline_dir = root / "baseline"
            baseline_dir.mkdir()
            old_workbook = baseline_dir / "resource_tree_baseline.xlsx"
            old_manifest = baseline_dir / "resource_tree_baseline_manifest.json"

            write_resource_tree_baseline(source)
            write_resource_tree_baseline(old_workbook)
            old_manifest.write_text(
                json.dumps({"sha256": hashlib.sha256(old_workbook.read_bytes()).hexdigest()}),
                encoding="utf-8",
            )

            manifest = promote_resource_tree_baseline(
                source,
                baseline_dir=baseline_dir,
                baseline_date="20260610",
                backup_timestamp="20260610T010203Z",
            )

            self.assertEqual("resource_tree_baseline", manifest["baseline_name"])
            self.assertTrue((baseline_dir / "resource_tree_baseline.xlsx").is_file())
            backup_dir = baseline_dir / "backup_resource_tree_20260610T010203Z"
            self.assertEqual(str(backup_dir), manifest["previous_baseline_backup"])
            self.assertTrue((backup_dir / "resource_tree_baseline_20260610T010203Z.xlsx").is_file())
            self.assertTrue((backup_dir / "resource_tree_baseline_manifest_20260610T010203Z.json").is_file())


if __name__ == "__main__":
    unittest.main()

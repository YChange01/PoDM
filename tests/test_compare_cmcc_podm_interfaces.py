from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from compare_cmcc_podm_interfaces import compare_cmcc_podm  # noqa: E402


class CompareCmccPodmInterfacesTest(unittest.TestCase):
    def test_compares_cmcc_request_response_against_podm_params(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            cmcc = root / "cmcc.yaml"
            toc = root / "cmcc.toc.yaml"
            podm = root / "podm.yaml"
            output = root / "compare.xlsx"
            cmcc.write_text(
                """
interfaces:
- section: 6.1.2.1
  title: 查询服务器资产
  method: GET
  uri: https://device_ip/redfish/v1/Systems/system_id
  params:
    request: [device_ip, system_id, auth_value]
    response: [Id, HostName]
- section: 6.1.9.99
  title: 未匹配接口
  method: POST
  uri: https://device_ip/redfish/v1/Missing
  params:
    request: [device_ip]
    response: []
""".strip()
                + "\n",
                encoding="utf-8",
            )
            toc.write_text(
                """
toc:
- section: 6.1.2
  level: 3
  title: 资产管理接口要求
- section: 6.1.2.1
  level: 4
  title: 查询服务器资产
""".strip()
                + "\n",
                encoding="utf-8",
            )
            podm.write_text(
                """
interfaces:
- section: 4.3.2
  title: 查询指定系统资源信息
  method: GET
  uri: /redfish/v1/Systems/{system_id}
  params:
    path: [system_id]
    header: [X-Auth-Token]
    body: []
    query: []
    response: [Id, AssetTag]
""".strip()
                + "\n",
                encoding="utf-8",
            )

            summary = compare_cmcc_podm(cmcc, podm, output, toc)
            wb = load_workbook(output, read_only=False, data_only=True)
            match_rows = list(wb["CMCC接口匹配"].iter_rows(min_row=2, values_only=True))
            param_ws = wb["共有接口参数对比"]
            param_rows = list(param_ws.iter_rows(min_row=2, values_only=True))
            request_row = next(row for row in param_rows if row[9] == "request")
            request_detail_rows = [row for row in param_rows if row[9] == "request" or row[15] or row[16]]
            merged_ranges = {str(item) for item in param_ws.merged_cells.ranges}

        self.assertEqual(summary["cmcc"], 2)
        self.assertEqual(summary["podm"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["unmatched"], 1)
        self.assertEqual(match_rows[0][0:4], ("6.1.2资产管理接口要求", "6.1.2.1查询服务器资产", None, "GET"))
        self.assertEqual(match_rows[0][7], "已匹配")
        self.assertEqual(match_rows[1][7], "未匹配")
        self.assertEqual(request_row[12], 2)
        self.assertEqual(request_row[13], 1)
        self.assertEqual(request_row[18], "device_ip")
        self.assertEqual(request_row[15], "device_ip")
        self.assertFalse(any("\n" in str(row[15] or "") for row in request_detail_rows))
        self.assertFalse(any("\n" in str(row[16] or "") for row in request_detail_rows))
        self.assertIn("A2:A6", merged_ranges)
        self.assertIn("J2:J4", merged_ranges)


if __name__ == "__main__":
    unittest.main()

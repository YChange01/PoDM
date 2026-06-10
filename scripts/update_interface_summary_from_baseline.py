#!/usr/bin/env python3
"""Update a reviewed Redfish interface summary from new interface-list YAML files."""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import defaultdict, deque
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _defaults import BMC_DOCX_NAME, OUTPUT_DIR, PODM_DOCX_NAME, REPO_ROOT  # noqa: E402


DEFAULT_BASELINE = REPO_ROOT / "baseline" / "reviewed_baseline.xlsx"
DEFAULT_BASELINE_MANIFEST = REPO_ROOT / "baseline" / "reviewed_baseline_manifest.json"


@dataclass(frozen=True)
class InterfaceItem:
    section: str
    title: str
    method: str
    uri: str
    index: int = 0


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_uri(uri: object) -> str:
    text = clean(uri)
    text = text.replace("新URL：", "")
    text = re.sub(r"^(GET|POST|PATCH|DELETE|PUT)\s+", "", text, flags=re.I)
    text = re.sub(r"^https?://[^/]+", "", text)
    text = text.replace(" ", "")
    text = text.replace("/Oem/Huawei/Public/", "/Oem/HuaweiPublic/")
    text = text.replace("/Oem/Huawei/Public", "/Oem/HuaweiPublic")
    text = text.replace("/Oem/Public/Huawei/", "/Oem/HuaweiPublic/")
    text = text.replace("/Oem/Huawei/", "/Oem/HuaweiPublic/")

    normalized_segments: list[str] = []
    for segment in text.split("/"):
        if not segment:
            continue
        if "?" in segment:
            base, query = segment.split("?", 1)
            query = re.sub(r"=[^&]+", "={id}", query)
            normalized_segments.append(f"{base}?{query}")
            continue
        if re.fullmatch(r"\{[^{}]+\}", segment):
            normalized_segments.append("{id}")
            continue
        if re.fullmatch(r"(?i)([a-z]+_)*[a-z]*id", segment):
            normalized_segments.append("{id}")
            continue
        normalized_segments.append(segment)
    return "/" + "/".join(normalized_segments)


def exact_key(item: InterfaceItem) -> tuple[str, str, str]:
    return (item.title, item.method.upper(), normalize_uri(item.uri))


def uri_key(item: InterfaceItem) -> tuple[str, str]:
    return (item.method.upper(), normalize_uri(item.uri))


def section_key(item: InterfaceItem) -> str:
    return item.section


def title_method_key(item: InterfaceItem) -> tuple[str, str]:
    return (item.title, item.method.upper())


def load_yaml_items(path: Path) -> list[InterfaceItem]:
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover - dependency is present in tests.
        raise SystemExit("读取 YAML 需要 PyYAML：python -m pip install pyyaml") from exc

    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    raw_items = data.get("interfaces") or []
    items: list[InterfaceItem] = []
    for raw in raw_items:
        items.append(
            InterfaceItem(
                section=clean(raw.get("section")),
                title=clean(raw.get("title")),
                method=clean(raw.get("method")),
                uri=clean(raw.get("uri")),
                index=int(raw.get("index") or len(items) + 1),
            )
        )
    return items


def item_from_row(row: dict[str, object], prefix: str = "") -> InterfaceItem:
    return InterfaceItem(
        section=clean(row.get(f"{prefix}section")),
        title=clean(row.get(f"{prefix}title")),
        method=clean(row.get(f"{prefix}method")),
        uri=clean(row.get(f"{prefix}uri")),
    )


def row_dicts(sheet) -> list[dict[str, object]]:
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers)})
    return rows


def read_baseline(path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    required = {"共有接口", "BMC独有", "PoDM独有"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(f"基线 workbook 缺少 sheet: {', '.join(sorted(missing))}")
    return row_dicts(wb["共有接口"]), row_dicts(wb["BMC独有"]), row_dicts(wb["PoDM独有"])


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def resolve_manifest_workbook(manifest: Path, value: object) -> Path | None:
    workbook = clean(value)
    if not workbook:
        return None
    raw_path = Path(workbook)
    candidates = [raw_path] if raw_path.is_absolute() else [REPO_ROOT / raw_path, manifest.parent / raw_path]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def validate_baseline_manifest(baseline: Path, manifest: Path) -> str:
    try:
        data = json.loads(manifest.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"baseline manifest 不是合法 JSON: {manifest}") from exc

    manifest_workbook = resolve_manifest_workbook(manifest, data.get("workbook"))
    if manifest_workbook and manifest_workbook.resolve() != baseline.resolve():
        raise SystemExit(
            "baseline manifest 指向的 workbook 与本次 baseline 不一致: "
            f"{manifest_workbook} != {baseline}"
        )

    expected = clean(data.get("sha256")).lower()
    if not expected:
        raise SystemExit(f"baseline manifest 缺少 sha256: {manifest}")

    actual = file_sha256(baseline)
    if actual.lower() != expected:
        raise SystemExit(f"baseline SHA256 校验失败: expected {expected}, actual {actual}")
    return actual


def pop_from_group(
    groups: dict[object, deque[int]],
    key: object,
    available: set[int],
) -> int | None:
    queue = groups.get(key)
    while queue and queue[0] not in available:
        queue.popleft()
    if queue:
        found = queue.popleft()
        available.remove(found)
        return found
    return None


def unique_group_match(
    groups: dict[object, list[int]],
    key: object,
    available: set[int],
) -> int | None:
    candidates = [idx for idx in groups.get(key, []) if idx in available]
    if len(candidates) == 1:
        available.remove(candidates[0])
        return candidates[0]
    return None


def build_match_indexes(items: list[InterfaceItem]) -> dict[str, object]:
    exact: dict[object, deque[int]] = defaultdict(deque)
    uri: dict[object, list[int]] = defaultdict(list)
    section: dict[object, list[int]] = defaultdict(list)
    title_method: dict[object, list[int]] = defaultdict(list)
    for index, item in enumerate(items):
        exact[exact_key(item)].append(index)
        uri[uri_key(item)].append(index)
        section[section_key(item)].append(index)
        title_method[title_method_key(item)].append(index)
    return {
        "exact": exact,
        "uri": uri,
        "section": section,
        "title_method": title_method,
    }


def match_baseline_item(
    baseline: InterfaceItem,
    new_items: list[InterfaceItem],
    indexes: dict[str, object],
    available: set[int],
) -> InterfaceItem | None:
    exact_groups = indexes["exact"]
    if isinstance(exact_groups, dict):
        idx = pop_from_group(exact_groups, exact_key(baseline), available)
        if idx is not None:
            return new_items[idx]

    for name, key in (
        ("uri", uri_key(baseline)),
        ("section", section_key(baseline)),
        ("title_method", title_method_key(baseline)),
    ):
        groups = indexes[name]
        if isinstance(groups, dict):
            idx = unique_group_match(groups, key, available)
            if idx is not None:
                return new_items[idx]
    return None


def field_change_note(side: str, baseline: InterfaceItem, current: InterfaceItem | None) -> str:
    if current is None:
        return f"{side}缺失"
    changes: list[str] = []
    if baseline.section != current.section:
        changes.append(f"section {baseline.section}->{current.section}")
    if baseline.title != current.title:
        changes.append(f"标题《{baseline.title}》->《{current.title}》")
    if baseline.method.upper() != current.method.upper():
        changes.append(f"method {baseline.method}->{current.method}")
    if normalize_uri(baseline.uri) != normalize_uri(current.uri):
        changes.append(f"URI {baseline.uri}->{current.uri}")
    elif baseline.uri != current.uri:
        changes.append(f"URI文本 {baseline.uri}->{current.uri}")
    if not changes:
        return ""
    prefix = f"{side}：" if side else ""
    return prefix + "，".join(changes)


def combine_notes(*parts: object) -> str:
    notes = [clean(part) for part in parts if clean(part)]
    return "；".join(notes)


def common_row(
    confidence: str,
    bmc: InterfaceItem,
    podm: InterfaceItem,
    note: str,
) -> list[str]:
    return [
        confidence,
        bmc.section,
        bmc.title,
        bmc.method,
        bmc.uri,
        podm.section,
        podm.title,
        podm.method,
        podm.uri,
        note,
    ]


def only_row(item: InterfaceItem, note: str) -> list[str]:
    return [item.section, item.title, item.method, item.uri, note]


def duplicate_rows(source: str, items: list[InterfaceItem]) -> list[list[object]]:
    groups: dict[tuple[str, str], list[InterfaceItem]] = defaultdict(list)
    for item in items:
        groups[uri_key(item)].append(item)
    rows: list[list[object]] = []
    group_id = 1
    for key, grouped in groups.items():
        if len(grouped) < 2:
            continue
        for item in grouped:
            rows.append([source, group_id, item.section, item.title, item.method, item.uri, key[0], key[1]])
        group_id += 1
    return rows


def exact_new_pairs(
    bmc_items: list[InterfaceItem],
    podm_items: list[InterfaceItem],
    available_bmc: set[int],
    available_podm: set[int],
) -> list[tuple[int, int]]:
    bmc_groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    podm_groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    for index in available_bmc:
        bmc_groups[uri_key(bmc_items[index])].append(index)
    for index in available_podm:
        podm_groups[uri_key(podm_items[index])].append(index)

    pairs: list[tuple[int, int]] = []
    for key, bmc_group in sorted(bmc_groups.items(), key=lambda item: min(item[1])):
        podm_group = podm_groups.get(key, [])
        if len(bmc_group) == 1 and len(podm_group) == 1:
            bmc_index, podm_index = bmc_group[0], podm_group[0]
            available_bmc.remove(bmc_index)
            available_podm.remove(podm_index)
            pairs.append((bmc_index, podm_index))
    return pairs


def new_report_summary() -> dict[str, dict[str, int]]:
    return {
        "common": {"rows": 0, "added": 0, "deleted": 0, "changed": 0},
        "bmc_only": {"rows": 0, "added": 0, "deleted": 0, "changed": 0},
        "podm_only": {"rows": 0, "added": 0, "deleted": 0, "changed": 0},
        "review_needed": {"common": 0, "bmc_only": 0, "podm_only": 0},
        "duplicates": {"rows": 0},
    }


def bump(summary: dict[str, dict[str, int]], sheet: str, status: str) -> None:
    summary[sheet][status] += 1


def update_summary_from_baseline(
    baseline: Path,
    bmc_yaml: Path,
    podm_yaml: Path,
    output: Path,
) -> dict[str, object]:
    baseline_common, baseline_bmc_only, baseline_podm_only = read_baseline(baseline)
    bmc_items = load_yaml_items(bmc_yaml)
    podm_items = load_yaml_items(podm_yaml)
    bmc_indexes = build_match_indexes(bmc_items)
    podm_indexes = build_match_indexes(podm_items)
    available_bmc = set(range(len(bmc_items)))
    available_podm = set(range(len(podm_items)))

    common_output: list[list[str]] = []
    bmc_only_output: list[list[str]] = []
    podm_only_output: list[list[str]] = []
    summary = new_report_summary()

    for row in baseline_common:
        bmc_base = item_from_row(row, "bmc_")
        podm_base = item_from_row(row, "podm_")
        bmc_current = match_baseline_item(bmc_base, bmc_items, bmc_indexes, available_bmc)
        podm_current = match_baseline_item(podm_base, podm_items, podm_indexes, available_podm)
        confidence = clean(row.get("confidence")) or "medium"
        baseline_note = row.get("复核备注") or row.get("备注") or row.get("匹配复核备注")
        if bmc_current is None or podm_current is None:
            note = combine_notes(
                baseline_note,
                "删除：基线共有配对在新接口文件中未完整出现（"
                + "，".join(
                    part
                    for part in (
                        "BMC缺失" if bmc_current is None else "",
                        "PoDM缺失" if podm_current is None else "",
                    )
                    if part
                )
                + "）",
            )
            common_output.append(common_row(confidence, bmc_current or bmc_base, podm_current or podm_base, note))
            bump(summary, "common", "deleted")
            continue
        change_note = combine_notes(
            field_change_note("BMC", bmc_base, bmc_current),
            field_change_note("PoDM", podm_base, podm_current),
        )
        note = combine_notes(baseline_note, f"变更：{change_note}" if change_note else "")
        common_output.append(common_row(confidence, bmc_current, podm_current, note))
        if change_note:
            bump(summary, "common", "changed")

    for row in baseline_bmc_only:
        base = item_from_row(row)
        current = match_baseline_item(base, bmc_items, bmc_indexes, available_bmc)
        baseline_note = row.get("差异备注") or row.get("备注")
        if current is None:
            note = combine_notes(baseline_note, "删除：基线BMC独有接口在新BMC接口文件中未出现")
            bmc_only_output.append(only_row(base, note))
            bump(summary, "bmc_only", "deleted")
            continue
        change_note = field_change_note("", base, current).lstrip("，")
        note = combine_notes(baseline_note, f"变更：{change_note}" if change_note else "")
        bmc_only_output.append(only_row(current, note))
        if change_note:
            bump(summary, "bmc_only", "changed")

    for row in baseline_podm_only:
        base = item_from_row(row)
        current = match_baseline_item(base, podm_items, podm_indexes, available_podm)
        baseline_note = row.get("差异备注") or row.get("备注")
        if current is None:
            note = combine_notes(baseline_note, "删除：基线PoDM独有接口在新PoDM接口文件中未出现")
            podm_only_output.append(only_row(base, note))
            bump(summary, "podm_only", "deleted")
            continue
        change_note = field_change_note("", base, current).lstrip("，")
        note = combine_notes(baseline_note, f"变更：{change_note}" if change_note else "")
        podm_only_output.append(only_row(current, note))
        if change_note:
            bump(summary, "podm_only", "changed")

    for bmc_index, podm_index in exact_new_pairs(bmc_items, podm_items, available_bmc, available_podm):
        bmc = bmc_items[bmc_index]
        podm = podm_items[podm_index]
        note = "新增：新版本出现的共有接口；依据 method + 归一化 URI 自动匹配"
        if bmc.title != podm.title:
            note += f"；标题不同：BMC《{bmc.title}》，PoDM《{podm.title}》"
        common_output.append(common_row("high", bmc, podm, note))
        bump(summary, "common", "added")

    for index in sorted(available_bmc, key=lambda idx: bmc_items[idx].index):
        bmc_only_output.append(
            only_row(
                bmc_items[index],
                "新增：新版本BMC接口，基线无对应项；需复核是否可与PoDM接口配对",
            )
        )
        bump(summary, "bmc_only", "added")
        bump(summary, "review_needed", "bmc_only")
    for index in sorted(available_podm, key=lambda idx: podm_items[idx].index):
        podm_only_output.append(
            only_row(
                podm_items[index],
                "新增：新版本PoDM接口，基线无对应项；需复核是否可与BMC接口配对",
            )
        )
        bump(summary, "podm_only", "added")
        bump(summary, "review_needed", "podm_only")

    duplicate_output = duplicate_rows("BMC", bmc_items) + duplicate_rows("PoDM", podm_items)
    summary["common"]["rows"] = len(common_output)
    summary["bmc_only"]["rows"] = len(bmc_only_output)
    summary["podm_only"]["rows"] = len(podm_only_output)
    summary["duplicates"]["rows"] = len(duplicate_output)

    write_workbook(output, common_output, bmc_only_output, podm_only_output, duplicate_output)
    return {
        "inputs": {
            "baseline": str(baseline),
            "bmc_yaml": str(bmc_yaml),
            "podm_yaml": str(podm_yaml),
        },
        "outputs": {"workbook": str(output)},
        "summary": summary,
    }


def write_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(1, column)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = ws.cell(row_index, column)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(
    output: Path,
    common_rows: list[list[object]],
    bmc_only_rows: list[list[object]],
    podm_only_rows: list[list[object]],
    duplicate_output: list[list[object]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "共有接口"
    write_sheet(
        ws,
        [
            "confidence",
            "bmc_section",
            "bmc_title",
            "bmc_method",
            "bmc_uri",
            "podm_section",
            "podm_title",
            "podm_method",
            "podm_uri",
            "复核备注",
        ],
        common_rows,
    )
    write_sheet(wb.create_sheet("BMC独有"), ["section", "title", "method", "uri", "差异备注"], bmc_only_rows)
    write_sheet(wb.create_sheet("PoDM独有"), ["section", "title", "method", "uri", "差异备注"], podm_only_rows)

    if duplicate_output:
        write_sheet(
            wb.create_sheet("YAML重复接口"),
            ["来源", "重复组", "section", "title", "method", "uri", "规范化method", "规范化uri"],
            duplicate_output,
        )

    widths = {
        "A": 14,
        "B": 18,
        "C": 34,
        "D": 12,
        "E": 80,
        "F": 18,
        "G": 34,
        "H": 12,
        "I": 80,
        "J": 72,
    }
    for ws in wb.worksheets:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="基于人工确认 Excel 基线和新的 interface-list YAML 生成带新增/删除/变更备注的 new_summary.xlsx"
    )
    parser.add_argument("date", nargs="?", help="新输出日期目录，例如 20260610")
    parser.add_argument("--baseline", type=Path, default=DEFAULT_BASELINE, help="人工确认基线 workbook")
    parser.add_argument(
        "--baseline-manifest",
        type=Path,
        help="基线 manifest JSON；使用默认 baseline 且 manifest 存在时会自动校验",
    )
    parser.add_argument("--skip-baseline-hash-check", action="store_true", help="跳过 baseline manifest SHA256 校验")
    parser.add_argument("--bmc-yaml", type=Path, help="新的 BMC interface-list YAML")
    parser.add_argument("--podm-yaml", type=Path, help="新的 PoDManager interface-list YAML")
    parser.add_argument("--output", type=Path, help="输出 xlsx；默认 output/<date>/analysis/new_summary.xlsx")
    parser.add_argument("--report", type=Path, help="输出 JSON 更新报告；默认与 new_summary.xlsx 同目录")
    return parser.parse_args(argv)


def default_yaml_paths(date: str) -> tuple[Path, Path]:
    out_dir = OUTPUT_DIR / date
    return (
        out_dir / f"{Path(BMC_DOCX_NAME).stem}.interface-list.yaml",
        out_dir / f"{Path(PODM_DOCX_NAME).stem}.interface-list.yaml",
    )


def default_manifest_for_baseline(baseline: Path) -> Path | None:
    if baseline.resolve() == DEFAULT_BASELINE.resolve() and DEFAULT_BASELINE_MANIFEST.is_file():
        return DEFAULT_BASELINE_MANIFEST
    return None


def write_report(path: Path, report: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main(argv: list[str] | None = None) -> None:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    if args.date:
        default_bmc, default_podm = default_yaml_paths(args.date)
        bmc_yaml = args.bmc_yaml or default_bmc
        podm_yaml = args.podm_yaml or default_podm
        output = args.output or (OUTPUT_DIR / args.date / "analysis" / "new_summary.xlsx")
    else:
        if not args.bmc_yaml or not args.podm_yaml or not args.output:
            raise SystemExit("未提供 date 时，必须显式指定 --bmc-yaml、--podm-yaml 和 --output")
        bmc_yaml = args.bmc_yaml
        podm_yaml = args.podm_yaml
        output = args.output
    report = args.report or (output.parent / "interface_update_report.json")

    for path in (args.baseline, bmc_yaml, podm_yaml):
        if not path.is_file():
            raise SystemExit(f"输入文件不存在: {path}")
    baseline_manifest = args.baseline_manifest or default_manifest_for_baseline(args.baseline)
    baseline_sha256 = ""
    if baseline_manifest and not args.skip_baseline_hash_check:
        if not baseline_manifest.is_file():
            raise SystemExit(f"baseline manifest 不存在: {baseline_manifest}")
        baseline_sha256 = validate_baseline_manifest(args.baseline, baseline_manifest)

    update_report = update_summary_from_baseline(args.baseline, bmc_yaml, podm_yaml, output)
    update_report["baseline_manifest"] = str(baseline_manifest) if baseline_manifest else ""
    update_report["baseline_sha256"] = baseline_sha256
    update_report["outputs"]["report"] = str(report)  # type: ignore[index]
    write_report(report, update_report)
    print(f"更新后的接口汇总已生成: {output}")
    print(f"接口更新报告已生成: {report}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Compare CMCC Redfish interface params against PoDManager interface params."""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _interface_compare import clean, normalize_uri  # noqa: E402
from _yaml_io import dedup_keep_order  # noqa: E402


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CMCC = REPO_ROOT / "output/20260609/cmcc.interface.yaml"
DEFAULT_CMCC_TOC = REPO_ROOT / "output/20260609/cmcc.toc.yaml"
DEFAULT_PODM = REPO_ROOT / "data/20260511/Atlas PoDManager 1.0.0 Redfish 接口参考.interfaces.yaml"
DEFAULT_OUTPUT = REPO_ROOT / "output/20260609/analysis/cmcc_podm_interface_param_compare.xlsx"
REQUEST_CATEGORIES = ("path", "header", "body", "query")
COMPARE_CATEGORIES = ("request", "response")
HEADER_ALIASES = {
    "x-auth-token": "auth_value",
    "authvalue": "auth_value",
    "authtoken": "auth_value",
    "content-type": "header_type",
    "contenttype": "header_type",
    "if-match": "ifmatch_value",
    "ifmatch": "ifmatch_value",
}


@dataclass(frozen=True)
class InterfaceRecord:
    index: int
    section: str
    title: str
    method: str
    uri: str
    params: dict[str, list[str]]


@dataclass(frozen=True)
class MatchResult:
    cmcc: InterfaceRecord
    podm: InterfaceRecord | None
    status: str
    mode: str
    candidate_count: int
    note: str = ""


@dataclass(frozen=True)
class TocEntry:
    section: str
    level: int
    title: str


def load_interfaces(path: Path) -> list[InterfaceRecord]:
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover - dependency is present in tests.
        raise SystemExit("读取 YAML 需要 PyYAML：python -m pip install pyyaml") from exc

    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    records: list[InterfaceRecord] = []
    for index, raw in enumerate(data.get("interfaces") or [], start=1):
        params = raw.get("params") or {}
        records.append(
            InterfaceRecord(
                index=int(raw.get("index") or index),
                section=clean(raw.get("section")),
                title=clean(raw.get("title")),
                method=clean(raw.get("method")).upper(),
                uri=clean(raw.get("uri")),
                params={key: [clean(item) for item in (value or [])] for key, value in params.items()},
            )
        )
    return records


def load_toc(path: Path | None) -> list[TocEntry]:
    if path is None or not path.exists():
        return []
    try:
        import yaml
    except ImportError as exc:  # pragma: no cover - dependency is present in tests.
        raise SystemExit("读取 YAML 需要 PyYAML：python -m pip install pyyaml") from exc

    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    entries: list[TocEntry] = []
    for raw in data.get("toc") or []:
        entries.append(
            TocEntry(
                section=clean(raw.get("section")),
                level=int(raw.get("level") or 0),
                title=clean(raw.get("title")),
            )
        )
    return entries


def ancestor_chapter(section: str, toc_entries: list[TocEntry], level: int) -> str:
    best: TocEntry | None = None
    for entry in toc_entries:
        if entry.level != level:
            continue
        if section == entry.section or section.startswith(f"{entry.section}."):
            if best is None or len(entry.section) > len(best.section):
                best = entry
    if best is None:
        return ""
    return f"{best.section} {best.title}".strip()


def match_key(item: InterfaceRecord) -> tuple[str, str]:
    return item.method.upper(), normalize_uri(item.uri).lower()


def title_score(left: str, right: str) -> float:
    return SequenceMatcher(None, left, right).ratio()


def match_interfaces(cmcc_items: list[InterfaceRecord], podm_items: list[InterfaceRecord]) -> list[MatchResult]:
    by_key: dict[tuple[str, str], list[InterfaceRecord]] = {}
    for item in podm_items:
        by_key.setdefault(match_key(item), []).append(item)

    results: list[MatchResult] = []
    for cmcc in cmcc_items:
        candidates = by_key.get(match_key(cmcc), [])
        if not candidates:
            results.append(MatchResult(cmcc, None, "未匹配", "", 0, "PoDManager 未找到 method + 归一化 URI 相同接口"))
            continue
        if len(candidates) == 1:
            results.append(MatchResult(cmcc, candidates[0], "已匹配", "归一化URI精确匹配", 1))
            continue
        best = max(candidates, key=lambda item: title_score(cmcc.title, item.title))
        results.append(
            MatchResult(
                cmcc,
                best,
                "已匹配",
                "归一化URI多候选，按标题相似度选择",
                len(candidates),
            )
        )
    return results


def request_params(item: InterfaceRecord) -> list[str]:
    if "request" in item.params:
        return item.params.get("request") or []
    out: list[str] = []
    for category in REQUEST_CATEGORIES:
        out.extend(item.params.get(category) or [])
    return out


def params_for(item: InterfaceRecord, category: str) -> list[str]:
    if category == "request":
        return request_params(item)
    return item.params.get(category) or []


def normalize_param_name(name: str) -> str:
    text = clean(name)
    text = text.strip("{}")
    text = text.replace("“", '"').replace("”", '"')
    text = re.sub(r"\s+", "", text)
    lowered = text.lower()
    lowered = HEADER_ALIASES.get(lowered, lowered)
    lowered = lowered.replace("/", ".")
    lowered = lowered.replace("oem.public", "oem.public")
    return lowered


def unique_param_names(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        normalized = normalize_param_name(item)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        out.append(item)
    return out


def param_diff(cmcc_params: list[str], podm_params: list[str]) -> dict[str, list[str]]:
    cmcc_unique = unique_param_names(cmcc_params)
    podm_unique = unique_param_names(podm_params)
    cmcc_map = {normalize_param_name(item): item for item in cmcc_unique}
    podm_map = {normalize_param_name(item): item for item in podm_unique}
    common_keys = [normalize_param_name(item) for item in cmcc_unique if normalize_param_name(item) in podm_map]
    cmcc_only = [item for item in cmcc_unique if normalize_param_name(item) not in podm_map]
    podm_only = [item for item in podm_unique if normalize_param_name(item) not in cmcc_map]
    return {
        "cmcc": cmcc_unique,
        "podm": podm_unique,
        "common": [cmcc_map[key] for key in dedup_keep_order(common_keys)],
        "cmcc_only": cmcc_only,
        "podm_only": podm_only,
    }


def join_values(values: list[str]) -> str:
    return "\n".join(values)


def count_summary(item: InterfaceRecord) -> str:
    request_count = len(unique_param_names(request_params(item)))
    response_count = len(unique_param_names(params_for(item, "response")))
    return f"request:{request_count}, response:{response_count}"


def write_workbook(results: list[MatchResult], output: Path, toc_entries: list[TocEntry] | None = None) -> None:
    toc_entries = toc_entries or []
    wb = Workbook()
    match_ws = wb.active
    match_ws.title = "CMCC接口匹配"
    match_ws.append(
        [
            "CMCC序号",
            "CMCC章节",
            "CMCC大章节",
            "CMCC标题",
            "方法",
            "CMCC URI",
            "CMCC归一化URI",
            "CMCC参数计数",
            "匹配状态",
            "匹配方式",
            "PODM章节",
            "PODM标题",
            "PODM URI",
            "PODM归一化URI",
            "PODM参数计数",
            "PODM候选数",
            "备注",
        ]
    )
    for result in results:
        podm = result.podm
        match_ws.append(
            [
                result.cmcc.index,
                result.cmcc.section,
                ancestor_chapter(result.cmcc.section, toc_entries, 3),
                result.cmcc.title,
                result.cmcc.method,
                result.cmcc.uri,
                normalize_uri(result.cmcc.uri),
                count_summary(result.cmcc),
                result.status,
                result.mode,
                podm.section if podm else "",
                podm.title if podm else "",
                podm.uri if podm else "",
                normalize_uri(podm.uri) if podm else "",
                count_summary(podm) if podm else "",
                result.candidate_count,
                result.note,
            ]
        )

    param_ws = wb.create_sheet("共有接口参数对比")
    param_ws.append(
        [
            "CMCC序号",
            "CMCC章节",
            "CMCC大章节",
            "CMCC标题",
            "方法",
            "CMCC URI",
            "PODM章节",
            "PODM标题",
            "PODM URI",
            "匹配方式",
            "参数类别",
            "CMCC参数数",
            "PODM参数数",
            "共同参数数",
            "CMCC独有数",
            "PODM独有数",
            "CMCC参数",
            "PODM参数",
            "共同参数",
            "CMCC独有",
            "PODM独有",
        ]
    )
    for result in results:
        if result.podm is None:
            continue
        start_row = param_ws.max_row + 1
        for category in COMPARE_CATEGORIES:
            diff = param_diff(params_for(result.cmcc, category), params_for(result.podm, category))
            rows = max(
                1,
                len(diff["cmcc"]),
                len(diff["podm"]),
                len(diff["common"]),
                len(diff["cmcc_only"]),
                len(diff["podm_only"]),
            )
            category_start = param_ws.max_row + 1
            for index in range(rows):
                param_ws.append(
                    [
                        result.cmcc.index,
                        result.cmcc.section,
                        ancestor_chapter(result.cmcc.section, toc_entries, 3),
                        result.cmcc.title,
                        result.cmcc.method,
                        result.cmcc.uri,
                        result.podm.section,
                        result.podm.title,
                        result.podm.uri,
                        result.mode,
                        category,
                        len(diff["cmcc"]),
                        len(diff["podm"]),
                        len(diff["common"]),
                        len(diff["cmcc_only"]),
                        len(diff["podm_only"]),
                        value_at(diff["cmcc"], index),
                        value_at(diff["podm"], index),
                        value_at(diff["common"], index),
                        value_at(diff["cmcc_only"], index),
                        value_at(diff["podm_only"], index),
                    ]
                )
            category_end = param_ws.max_row
            if category_end > category_start:
                for col in range(11, 17):
                    param_ws.merge_cells(
                        start_row=category_start,
                        start_column=col,
                        end_row=category_end,
                        end_column=col,
                    )
        end_row = param_ws.max_row
        if end_row > start_row:
            for col in range(1, 10):
                param_ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

    style_workbook(wb)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def value_at(values: list[str], index: int) -> str:
    return values[index] if index < len(values) else ""


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                parts = str(cell.value).splitlines() or [""]
                max_len = max(max_len, max(len(part) for part in parts))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def compare_cmcc_podm(
    cmcc_yaml: Path,
    podm_yaml: Path,
    output: Path,
    cmcc_toc: Path | None = DEFAULT_CMCC_TOC,
) -> dict[str, int]:
    cmcc_items = load_interfaces(cmcc_yaml)
    podm_items = load_interfaces(podm_yaml)
    toc_entries = load_toc(cmcc_toc)
    results = match_interfaces(cmcc_items, podm_items)
    write_workbook(results, output, toc_entries)
    matched = sum(1 for result in results if result.podm is not None)
    return {
        "cmcc": len(cmcc_items),
        "podm": len(podm_items),
        "matched": matched,
        "unmatched": len(results) - matched,
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="对比 CMCC 与 PoDManager 接口及参数，输出 Excel。")
    parser.add_argument("--cmcc-yaml", type=Path, default=DEFAULT_CMCC)
    parser.add_argument("--cmcc-toc", type=Path, default=DEFAULT_CMCC_TOC)
    parser.add_argument("--podm-yaml", type=Path, default=DEFAULT_PODM)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    for path in (args.cmcc_yaml, args.podm_yaml):
        if not path.exists():
            raise SystemExit(f"输入文件不存在: {path}")
    summary = compare_cmcc_podm(args.cmcc_yaml, args.podm_yaml, args.output, args.cmcc_toc)
    print(
        "CMCC vs PoDManager 对比完成："
        f"CMCC {summary['cmcc']} 条 / PoDManager {summary['podm']} 条 / "
        f"已匹配 {summary['matched']} 条 / 未匹配 {summary['unmatched']} 条"
    )
    print(f"Excel: {args.output}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Promote a reviewed interface summary workbook to the repository baseline."""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _defaults import OUTPUT_DIR, REPO_ROOT  # noqa: E402


DEFAULT_BASELINE_DIR = REPO_ROOT / "baseline"
DEFAULT_BASELINE_WORKBOOK = "reviewed_baseline.xlsx"
DEFAULT_BASELINE_MANIFEST = "reviewed_baseline_manifest.json"
NOTE_COLUMNS = {"复核备注", "差异备注", "备注", "匹配复核备注"}
REQUIRED_HEADERS = {
    "共有接口": [
        "confidence",
        "bmc_section",
        "bmc_title",
        "bmc_method",
        "bmc_uri",
        "podm_section",
        "podm_title",
        "podm_method",
        "podm_uri",
        "复核备注",
    ],
    "BMC独有": ["section", "title", "method", "uri", "差异备注"],
    "PoDM独有": ["section", "title", "method", "uri", "差异备注"],
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def repo_relative(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(REPO_ROOT.resolve()))
    except ValueError:
        return str(path)


def current_backup_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def unique_backup_dir(baseline_dir: Path, timestamp: str) -> Path:
    base = baseline_dir / f"backup_{timestamp}"
    if not base.exists():
        return base
    suffix = 1
    while True:
        candidate = baseline_dir / f"backup_{timestamp}_{suffix}"
        if not candidate.exists():
            return candidate
        suffix += 1


def timestamped_backup_name(path: Path, timestamp: str) -> str:
    return f"{path.stem}_{timestamp}{path.suffix}"


def archive_existing_baseline(baseline_dir: Path, timestamp: str) -> Path | None:
    existing_paths = [
        path
        for path in (
            baseline_dir / DEFAULT_BASELINE_WORKBOOK,
            baseline_dir / DEFAULT_BASELINE_MANIFEST,
        )
        if path.exists()
    ]
    if not existing_paths:
        return None

    backup_dir = unique_backup_dir(baseline_dir, timestamp)
    backup_dir.mkdir(parents=True, exist_ok=False)
    for path in existing_paths:
        shutil.copy2(path, backup_dir / timestamped_backup_name(path, timestamp))
    return backup_dir


def sheet_summary(workbook: Path) -> dict[str, dict[str, object]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    summary: dict[str, dict[str, object]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_cells = next(ws.iter_rows(min_row=1, max_row=1), ())
        headers = [clean(cell.value) for cell in header_cells]
        note_counts: dict[str, int] = {header: 0 for header in headers if header in NOTE_COLUMNS}
        for values in ws.iter_rows(min_row=2, values_only=True):
            for index, header in enumerate(headers):
                if header in note_counts and index < len(values) and clean(values[index]):
                    note_counts[header] += 1
        summary[sheet_name] = {
            "rows": max(ws.max_row - 1, 0),
            "columns": ws.max_column,
            "headers": headers,
            "note_counts": note_counts,
        }
    return summary


def validate_reviewed_workbook(workbook: Path) -> None:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    missing_sheets = set(REQUIRED_HEADERS) - set(wb.sheetnames)
    if missing_sheets:
        raise SystemExit(f"待提升 workbook 缺少 sheet: {', '.join(sorted(missing_sheets))}")

    for sheet_name, required_headers in REQUIRED_HEADERS.items():
        ws = wb[sheet_name]
        header_cells = next(ws.iter_rows(min_row=1, max_row=1), ())
        headers = [clean(cell.value) for cell in header_cells]
        if headers[: len(required_headers)] != required_headers:
            raise SystemExit(
                f"待提升 workbook 的 {sheet_name} 表头不符合要求: "
                f"expected {required_headers}, actual {headers}"
            )


def promote_reviewed_baseline(
    source_workbook: Path,
    baseline_dir: Path = DEFAULT_BASELINE_DIR,
    baseline_date: str = "",
    backup_timestamp: str | None = None,
) -> dict[str, object]:
    if not source_workbook.is_file():
        raise SystemExit(f"待提升的 workbook 不存在: {source_workbook}")
    validate_reviewed_workbook(source_workbook)

    baseline_dir.mkdir(parents=True, exist_ok=True)
    target_workbook = baseline_dir / DEFAULT_BASELINE_WORKBOOK
    backup_dir = None
    if source_workbook.resolve() != target_workbook.resolve():
        backup_dir = archive_existing_baseline(
            baseline_dir,
            backup_timestamp or current_backup_timestamp(),
        )
        shutil.copy2(source_workbook, target_workbook)

    manifest = {
        "baseline_name": "reviewed_baseline",
        "baseline_date": baseline_date,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "workbook": repo_relative(target_workbook),
        "sha256": file_sha256(target_workbook),
        "source_workbook": repo_relative(source_workbook),
        "previous_baseline_backup": repo_relative(backup_dir) if backup_dir else "",
        "sheets": sheet_summary(target_workbook),
    }
    manifest_path = baseline_dir / DEFAULT_BASELINE_MANIFEST
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="把人工复核后的 new_summary.xlsx 提升为 baseline/reviewed_baseline.xlsx，并重新生成 manifest"
    )
    parser.add_argument(
        "date",
        nargs="?",
        help="输出日期目录，例如 20260610；默认读取 output/<date>/analysis/new_summary.xlsx",
    )
    parser.add_argument("--source", type=Path, help="人工复核后的 summary workbook")
    parser.add_argument("--baseline-dir", type=Path, default=DEFAULT_BASELINE_DIR, help="baseline 输出目录")
    parser.add_argument("--baseline-date", help="写入 manifest 的 baseline_date；默认使用 date")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    if args.source:
        source = args.source
    elif args.date:
        source = OUTPUT_DIR / args.date / "analysis" / "new_summary.xlsx"
    else:
        raise SystemExit("必须提供 date 或 --source")

    baseline_date = args.baseline_date or args.date or ""
    manifest = promote_reviewed_baseline(source, args.baseline_dir, baseline_date=baseline_date)
    print(f"已更新基线 workbook: {args.baseline_dir / DEFAULT_BASELINE_WORKBOOK}")
    print(f"已更新基线 manifest: {args.baseline_dir / DEFAULT_BASELINE_MANIFEST}")
    if manifest.get("previous_baseline_backup"):
        print(f"旧基线已归档: {manifest['previous_baseline_backup']}")
    print(f"baseline SHA256: {manifest['sha256']}")


if __name__ == "__main__":
    main()

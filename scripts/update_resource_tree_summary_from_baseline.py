#!/usr/bin/env python3
"""Compare resource-tree interfaces with PoDManager interface-list YAML."""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _defaults import OUTPUT_DIR, PODM_DOCX_NAME, REPO_ROOT  # noqa: E402
from update_interface_summary_from_baseline import (  # noqa: E402
    InterfaceItem,
    build_match_indexes,
    clean,
    combine_notes,
    exact_new_pairs,
    field_change_note,
    item_from_row,
    load_yaml_items,
    match_baseline_item,
    normalize_uri,
    uri_key,
)

DEFAULT_BASELINE = REPO_ROOT / "baseline" / "resource_tree_baseline.xlsx"
DEFAULT_BASELINE_MANIFEST = REPO_ROOT / "baseline" / "resource_tree_baseline_manifest.json"

COMMON_HEADERS = [
    "confidence",
    "resource_section",
    "resource_title",
    "resource_method",
    "resource_uri",
    "podm_section",
    "podm_title",
    "podm_method",
    "podm_uri",
    "复核备注",
]
ONLY_HEADERS = ["section", "title", "method", "uri", "差异备注"]


def resource_tree_yaml_for_date(date: str) -> Path:
    return OUTPUT_DIR / date / f"{Path(PODM_DOCX_NAME).stem}.resource-tree.interface-list.yaml"


def podm_yaml_for_date(date: str) -> Path:
    return OUTPUT_DIR / date / f"{Path(PODM_DOCX_NAME).stem}.interface-list.yaml"


def row_dicts(sheet) -> list[dict[str, object]]:
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers)})
    return rows


def read_baseline(path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    required = {"共有接口", "资源树独有", "PoDM独有"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise SystemExit(f"资源树基线 workbook 缺少 sheet: {', '.join(sorted(missing))}")
    try:
        return row_dicts(wb["共有接口"]), row_dicts(wb["资源树独有"]), row_dicts(wb["PoDM独有"])
    finally:
        wb.close()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_baseline_manifest(baseline: Path, manifest: Path) -> str:
    try:
        data = json.loads(manifest.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"资源树 baseline manifest 不是合法 JSON: {manifest}") from exc

    expected = clean(data.get("sha256")).lower()
    if not expected:
        raise SystemExit(f"资源树 baseline manifest 缺少 sha256: {manifest}")
    actual = file_sha256(baseline)
    if actual.lower() != expected:
        raise SystemExit(f"资源树 baseline SHA256 校验失败: expected {expected}, actual {actual}")
    workbook = clean(data.get("workbook"))
    if workbook:
        workbook_path = Path(workbook)
        if not workbook_path.is_absolute():
            workbook_path = REPO_ROOT / workbook_path
        if workbook_path.exists() and workbook_path.resolve() != baseline.resolve():
            raise SystemExit(f"资源树 baseline manifest 指向的 workbook 不一致: {workbook_path} != {baseline}")
    return actual


def common_row(confidence: str, resource: InterfaceItem, podm: InterfaceItem, note: str) -> list[str]:
    return [
        confidence,
        resource.section,
        resource.title,
        resource.method,
        resource.uri,
        podm.section,
        podm.title,
        podm.method,
        podm.uri,
        note,
    ]


def only_row(item: InterfaceItem, note: str) -> list[str]:
    return [item.section, item.title, item.method, item.uri, note]


def new_report_summary() -> dict[str, dict[str, int]]:
    return {
        "common": {"rows": 0, "added": 0, "deleted": 0, "changed": 0},
        "resource_tree_only": {"rows": 0, "added": 0, "deleted": 0, "changed": 0},
        "podm_only": {"rows": 0, "added": 0, "deleted": 0, "changed": 0},
        "review_needed": {"resource_tree_only": 0, "podm_only": 0},
        "duplicates": {"rows": 0},
    }


def bump(summary: dict[str, dict[str, int]], sheet: str, status: str) -> None:
    summary[sheet][status] += 1


def duplicate_rows(source: str, items: list[InterfaceItem]) -> list[list[object]]:
    groups: dict[tuple[str, str], list[InterfaceItem]] = defaultdict(list)
    for item in items:
        groups[uri_key(item)].append(item)

    rows: list[list[object]] = []
    group_id = 1
    for grouped in groups.values():
        if len(grouped) < 2:
            continue
        for item in grouped:
            rows.append([source, group_id, item.section, item.title, item.method, item.uri])
        group_id += 1
    return rows


def direct_compare(resource_items: list[InterfaceItem], podm_items: list[InterfaceItem]) -> tuple[list[list[str]], list[list[str]], list[list[str]], dict[str, dict[str, int]]]:
    resource_indexes = build_match_indexes(resource_items)
    podm_indexes = build_match_indexes(podm_items)
    available_resource = set(range(len(resource_items)))
    available_podm = set(range(len(podm_items)))
    summary = new_report_summary()
    common_output: list[list[str]] = []

    for resource_index, podm_index in exact_new_pairs(resource_items, podm_items, available_resource, available_podm):
        common_output.append(common_row("high", resource_items[resource_index], podm_items[podm_index], ""))

    resource_only_output = [only_row(resource_items[index], "") for index in sorted(available_resource, key=lambda idx: resource_items[idx].index)]
    podm_only_output = [only_row(podm_items[index], "") for index in sorted(available_podm, key=lambda idx: podm_items[idx].index)]
    summary["common"]["rows"] = len(common_output)
    summary["resource_tree_only"]["rows"] = len(resource_only_output)
    summary["podm_only"]["rows"] = len(podm_only_output)
    summary["duplicates"]["rows"] = len(duplicate_rows("资源树", resource_items) + duplicate_rows("PoDM", podm_items))
    return common_output, resource_only_output, podm_only_output, summary


def baseline_compare(
    baseline: Path,
    resource_items: list[InterfaceItem],
    podm_items: list[InterfaceItem],
) -> tuple[list[list[str]], list[list[str]], list[list[str]], dict[str, dict[str, int]]]:
    baseline_common, baseline_resource_only, baseline_podm_only = read_baseline(baseline)
    resource_indexes = build_match_indexes(resource_items)
    podm_indexes = build_match_indexes(podm_items)
    available_resource = set(range(len(resource_items)))
    available_podm = set(range(len(podm_items)))

    common_output: list[list[str]] = []
    resource_only_output: list[list[str]] = []
    podm_only_output: list[list[str]] = []
    summary = new_report_summary()

    for row in baseline_common:
        resource_base = item_from_row(row, "resource_")
        podm_base = item_from_row(row, "podm_")
        resource_current = match_baseline_item(resource_base, resource_items, resource_indexes, available_resource)
        podm_current = match_baseline_item(podm_base, podm_items, podm_indexes, available_podm)
        confidence = clean(row.get("confidence")) or "medium"
        baseline_note = row.get("复核备注") or row.get("备注") or row.get("匹配复核备注")
        if resource_current is None or podm_current is None:
            note = combine_notes(
                baseline_note,
                "删除：基线共有配对在新接口文件中未完整出现（"
                + "，".join(
                    part
                    for part in (
                        "资源树缺失" if resource_current is None else "",
                        "PoDM缺失" if podm_current is None else "",
                    )
                    if part
                )
                + "）",
            )
            common_output.append(common_row(confidence, resource_current or resource_base, podm_current or podm_base, note))
            bump(summary, "common", "deleted")
            continue
        change_note = combine_notes(
            field_change_note("资源树", resource_base, resource_current),
            field_change_note("PoDM", podm_base, podm_current),
        )
        note = combine_notes(baseline_note, f"变更：{change_note}" if change_note else "")
        common_output.append(common_row(confidence, resource_current, podm_current, note))
        if change_note:
            bump(summary, "common", "changed")

    for row in baseline_resource_only:
        base = item_from_row(row)
        current = match_baseline_item(base, resource_items, resource_indexes, available_resource)
        baseline_note = row.get("差异备注") or row.get("备注")
        if current is None:
            resource_only_output.append(only_row(base, combine_notes(baseline_note, "删除：基线资源树独有接口在新资源树中未出现")))
            bump(summary, "resource_tree_only", "deleted")
            continue
        change_note = field_change_note("", base, current).lstrip("，")
        resource_only_output.append(only_row(current, combine_notes(baseline_note, f"变更：{change_note}" if change_note else "")))
        if change_note:
            bump(summary, "resource_tree_only", "changed")

    for row in baseline_podm_only:
        base = item_from_row(row)
        current = match_baseline_item(base, podm_items, podm_indexes, available_podm)
        baseline_note = row.get("差异备注") or row.get("备注")
        if current is None:
            podm_only_output.append(only_row(base, combine_notes(baseline_note, "删除：基线PoDM独有接口在新PoDM接口文件中未出现")))
            bump(summary, "podm_only", "deleted")
            continue
        change_note = field_change_note("", base, current).lstrip("，")
        podm_only_output.append(only_row(current, combine_notes(baseline_note, f"变更：{change_note}" if change_note else "")))
        if change_note:
            bump(summary, "podm_only", "changed")

    for resource_index, podm_index in exact_new_pairs(resource_items, podm_items, available_resource, available_podm):
        resource = resource_items[resource_index]
        podm = podm_items[podm_index]
        common_output.append(
            common_row(
                "high",
                resource,
                podm,
                "新增：新版本资源树和PoDM接口文件均出现；依据 method + 归一化 URI 自动匹配",
            )
        )
        bump(summary, "common", "added")

    for index in sorted(available_resource, key=lambda idx: resource_items[idx].index):
        resource_only_output.append(
            only_row(resource_items[index], "新增：新版本资源树接口，基线无对应项；需复核PoDM接口文件是否缺少")
        )
        bump(summary, "resource_tree_only", "added")
        bump(summary, "review_needed", "resource_tree_only")

    for index in sorted(available_podm, key=lambda idx: podm_items[idx].index):
        podm_only_output.append(
            only_row(podm_items[index], "新增：新版本PoDM接口，基线无对应资源树接口；需复核资源树是否缺少")
        )
        bump(summary, "podm_only", "added")
        bump(summary, "review_needed", "podm_only")

    summary["common"]["rows"] = len(common_output)
    summary["resource_tree_only"]["rows"] = len(resource_only_output)
    summary["podm_only"]["rows"] = len(podm_only_output)
    return common_output, resource_only_output, podm_only_output, summary


def write_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(1, column)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = ws.cell(row_index, column)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(
    output: Path,
    common_rows: list[list[object]],
    resource_only_rows: list[list[object]],
    podm_only_rows: list[list[object]],
    duplicate_output: list[list[object]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "共有接口"
    write_sheet(ws, COMMON_HEADERS, common_rows)
    write_sheet(wb.create_sheet("资源树独有"), ONLY_HEADERS, resource_only_rows)
    write_sheet(wb.create_sheet("PoDM独有"), ONLY_HEADERS, podm_only_rows)
    if duplicate_output:
        write_sheet(wb.create_sheet("重复接口"), ["来源", "重复组", "section", "title", "method", "uri"], duplicate_output)

    widths = {"A": 16, "B": 22, "C": 58, "D": 12, "E": 84, "F": 18, "G": 34, "H": 12, "I": 84, "J": 72}
    for ws in wb.worksheets:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def update_resource_tree_summary(
    baseline: Path | None,
    resource_tree_yaml: Path,
    podm_yaml: Path,
    output: Path,
) -> dict[str, object]:
    resource_items = load_yaml_items(resource_tree_yaml)
    podm_items = load_yaml_items(podm_yaml)
    duplicate_output = duplicate_rows("资源树", resource_items) + duplicate_rows("PoDM", podm_items)

    if baseline and baseline.is_file():
        common_rows, resource_only_rows, podm_only_rows, summary = baseline_compare(baseline, resource_items, podm_items)
    else:
        common_rows, resource_only_rows, podm_only_rows, summary = direct_compare(resource_items, podm_items)
    summary["duplicates"]["rows"] = len(duplicate_output)

    write_workbook(output, common_rows, resource_only_rows, podm_only_rows, duplicate_output)
    return {
        "inputs": {
            "baseline": str(baseline) if baseline else "",
            "resource_tree_yaml": str(resource_tree_yaml),
            "podm_yaml": str(podm_yaml),
        },
        "outputs": {"workbook": str(output)},
        "summary": summary,
    }


def default_manifest_for_baseline(baseline: Path) -> Path | None:
    if baseline.resolve() == DEFAULT_BASELINE.resolve() and DEFAULT_BASELINE_MANIFEST.is_file():
        return DEFAULT_BASELINE_MANIFEST
    return None


def write_report(path: Path, report: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="对比资源树接口和 PoDManager 接口清单，生成资源树差异 workbook")
    parser.add_argument("date", nargs="?", help="新输出日期目录，例如 20260610")
    parser.add_argument("--baseline", type=Path, default=DEFAULT_BASELINE, help="资源树人工确认基线 workbook")
    parser.add_argument("--baseline-manifest", type=Path, help="资源树基线 manifest JSON")
    parser.add_argument("--skip-baseline-hash-check", action="store_true", help="跳过 baseline manifest SHA256 校验")
    parser.add_argument("--resource-tree-yaml", type=Path, help="新的资源树 interface-list YAML")
    parser.add_argument("--podm-yaml", type=Path, help="新的 PoDManager interface-list YAML")
    parser.add_argument("--output", type=Path, help="输出 xlsx；默认 output/<date>/analysis/resource_tree_summary.xlsx")
    parser.add_argument("--report", type=Path, help="输出 JSON 更新报告；默认与 workbook 同目录")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    if args.date:
        resource_tree_yaml = args.resource_tree_yaml or resource_tree_yaml_for_date(args.date)
        podm_yaml = args.podm_yaml or podm_yaml_for_date(args.date)
        output = args.output or (OUTPUT_DIR / args.date / "analysis" / "resource_tree_summary.xlsx")
    else:
        if not args.resource_tree_yaml or not args.podm_yaml or not args.output:
            raise SystemExit("未提供 date 时，必须显式指定 --resource-tree-yaml、--podm-yaml 和 --output")
        resource_tree_yaml = args.resource_tree_yaml
        podm_yaml = args.podm_yaml
        output = args.output
    report = args.report or (output.parent / "resource_tree_update_report.json")

    for path in (resource_tree_yaml, podm_yaml):
        if not path.is_file():
            raise SystemExit(f"输入文件不存在: {path}")

    baseline = args.baseline if args.baseline and args.baseline.is_file() else None
    baseline_manifest = args.baseline_manifest or (default_manifest_for_baseline(args.baseline) if baseline else None)
    baseline_sha256 = ""
    if baseline and baseline_manifest and not args.skip_baseline_hash_check:
        if not baseline_manifest.is_file():
            raise SystemExit(f"资源树 baseline manifest 不存在: {baseline_manifest}")
        baseline_sha256 = validate_baseline_manifest(baseline, baseline_manifest)

    update_report = update_resource_tree_summary(baseline, resource_tree_yaml, podm_yaml, output)
    update_report["baseline_manifest"] = str(baseline_manifest) if baseline_manifest else ""
    update_report["baseline_sha256"] = baseline_sha256
    update_report["outputs"]["report"] = str(report)  # type: ignore[index]
    write_report(report, update_report)
    print(f"资源树接口汇总已生成: {output}")
    print(f"资源树接口更新报告已生成: {report}")


if __name__ == "__main__":
    main()

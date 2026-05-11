#!/usr/bin/env python3
"""Extract Redfish interfaces plus parameter names/types from Word documents.

This script is intentionally independent from the existing extractors. It reads
`.docx` or already-flattened `.txt` Word exports and emits a structured file that
is suitable for later BMC-vs-PoDM parameter comparison.

Examples:
    python3 scripts/extract_word_interface_params.py --profile podm \
      "data/20260507/Atlas PoDManager 1.0.0 Redfish 接口参考.docx" \
      -o "output/20260507/podm.word.interface-params.yaml"

    python3 scripts/extract_word_interface_params.py --profile bmc \
      --match-workbook output/20260507/analysis/interface_match_llm_summary.xlsx \
      --match-side bmc \
      "data/20260507/华为服务器 iBMC300 Redfish 接口说明.docx" \
      -o "output/20260507/bmc.common.word.interface-params.yaml"
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable


CATEGORIES = ("path", "header", "body", "query", "response")
W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

HEADING_RE = re.compile(r"^(\d+(?:\.\d+)+)[\s\t]+([^\s/{][^{}]{0,100})$")
TEXT_NUM_PREFIX_RE = re.compile(r"^(\d+(?:\.\d+)*)\s+(.*)$")
TRAILING_PAGENO_RE = re.compile(r"(\d{1,4})$")

PODM_MARKERS = (
    "接口功能",
    "接口约束",
    "调用方法",
    "URI",
    "请求参数",
    "请求示例",
    "响应参数",
    "响应示例",
    "返回值",
    "样例",
)
BMC_MARKERS = ("命令功能", "命令格式", "参数说明", "使用指南", "使用实例", "输出说明")

PODM_MARKER_RE = re.compile(
    r"^(?:\S{1,5}板)?("
    + "|".join(re.escape(marker) for marker in PODM_MARKERS)
    + r")\s*\d*$"
)
BMC_MARKER_RE = re.compile(
    r"^(" + "|".join(re.escape(marker) for marker in BMC_MARKERS) + r")\s*\d*$"
)
TABLE_TITLE_RE = re.compile(
    r"^(?:表\s*[\d\-.]+\s+)?(?:Path|Header|Body|Query|Response)?参数列表\s*$",
    re.IGNORECASE,
)
ANY_TABLE_HEADER_RE = re.compile(r"^表\s*[\d\-.]+\s+\S")
STATUS_CODE_PROSE_RE = re.compile(r"^返回状态码为\d+")

METHOD_LINE_RE = re.compile(r"^操作类型\s*[:：]\s*(\w+)\s*$")
URL_LINE_RE = re.compile(
    r"^(?:[一-鿿]{1,8}\s*[:：]\s*)?"
    r"[一-鿿]{0,3}URL\s*[:：]\s*"
    r"(.*)$"
)
HEADER_LINE_RE = re.compile(r"^([A-Za-z][\w-]*)\s*:\s*\S+")
JSON_KEY_RE = re.compile(r'"([@A-Za-z_][\w@.\-]*)"\s*:')
NAME_LIKE_RE = re.compile(r"^[$@A-Za-z_][A-Za-z0-9_$@./\-]*$")
BRACED_PLACEHOLDER_RE = re.compile(r"\{+([^{}]+)\}+")

TYPE_VALUES = {
    "string",
    "str",
    "integer",
    "int",
    "int32",
    "int64",
    "uint",
    "uint32",
    "uint64",
    "long",
    "short",
    "number",
    "float",
    "double",
    "decimal",
    "boolean",
    "bool",
    "array",
    "object",
    "enum",
    "null",
    "uri",
    "url",
    "date",
    "datetime",
    "date-time",
    "time",
    "timestamp",
    "ipv4address",
    "ipv6address",
    "macaddress",
    "guid",
    "uuid",
    "password",
    "byte",
    "bytes",
    "map",
    "dict",
    "list",
    "array[object]",
    "array[string]",
    "字符串",
    "字符",
    "字符型",
    "整数",
    "整型",
    "数字",
    "布尔",
    "布尔值",
    "布尔型",
    "对象",
    "数组",
    "列表",
    "枚举",
    "枚举型",
    "浮点",
    "浮点数",
    "浮点型",
    "时间",
    "时间戳",
    "日期",
    "字节",
    "属性",
    "自定义属性",
    "集合",
    "映射",
    "字典",
    "-",
    "–",
}
TYPE_SUFFIX_RE = re.compile(r"^.{0,24}(列表|数组|对象|集合|属性|字典|映射|型)$")
PATH_KEYWORDS = {"redfish", "v1", "actions", "oem", "huawei", "public"}
IGNORED_PARAM_NAMES = {"device_ip"}


@dataclass
class ParamRecord:
    name: str
    type: str = ""
    required: str = ""
    description: str = ""
    value: str = ""
    inferred: bool = False


@dataclass
class InterfaceRecord:
    index: int
    section: str
    title: str
    method: str
    uri: str
    params: dict[str, list[ParamRecord]] = field(
        default_factory=lambda: {category: [] for category in CATEGORIES}
    )


def read_docx(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        with archive.open("word/document.xml") as handle:
            tree = ET.parse(handle)

    body = tree.getroot().find(f"{W_NS}body")
    if body is None:
        return ""

    out: list[str] = []
    counters = [0] * 9

    for child in body:
        if child.tag == f"{W_NS}p":
            text = para_text(child)
            style = pstyle_value(child)
            if is_toc_style(style):
                continue
            level = heading_level(style)
            if level is None:
                out.append(text)
                continue
            stripped = text.strip()
            match = TEXT_NUM_PREFIX_RE.match(stripped)
            if match:
                nums = [int(part) for part in match.group(1).split(".")]
                for i, value in enumerate(nums):
                    if i < len(counters):
                        counters[i] = value
                for i in range(len(nums), len(counters)):
                    counters[i] = 0
                out.append(text)
            elif stripped:
                counters[level - 1] += 1
                for i in range(level, len(counters)):
                    counters[i] = 0
                number = ".".join(str(value) for value in counters[:level])
                out.append(f"{number} {stripped}")
            else:
                out.append(text)
        elif child.tag == f"{W_NS}tbl":
            out.extend(table_rows(child))

    return "\n".join(out)


def read_source(path: Path) -> str:
    if path.suffix.lower() == ".docx":
        text = read_docx(path)
    else:
        text = path.read_text(encoding="utf-8-sig")
    return text.replace("　", " ").replace("\xa0", " ")


def para_text(p_elem: ET.Element) -> str:
    return "".join((text.text or "") for text in p_elem.iter(f"{W_NS}t"))


def table_rows(tbl_elem: ET.Element) -> list[str]:
    rows: list[str] = []
    for row in tbl_elem.iter(f"{W_NS}tr"):
        cells = [
            "".join((text.text or "") for text in cell.iter(f"{W_NS}t"))
            for cell in row.iter(f"{W_NS}tc")
        ]
        rows.append("\t".join(cells))
    return rows


def pstyle_value(p_elem: ET.Element) -> str:
    style = p_elem.find(f"{W_NS}pPr/{W_NS}pStyle")
    return (style.get(f"{W_NS}val") if style is not None else "") or ""


def normalized_style(style: str) -> str:
    return style.strip().lower().replace(" ", "").replace("-", "").replace("_", "")


def heading_level(style: str) -> int | None:
    normalized = normalized_style(style)
    for level in range(1, 10):
        if normalized in {f"heading{level}", f"标题{level}", f"hdg{level}"}:
            return level
    return None


def is_toc_style(style: str) -> bool:
    normalized = normalized_style(style)
    return normalized.startswith("toc") or normalized.startswith("目录")


def strip_trailing_pageno(title: str) -> str:
    match = TRAILING_PAGENO_RE.search(title)
    if match and match.start() > 0 and not title[match.start() - 1].isdigit():
        return title[: match.start()]
    return title


def split_sections(text: str) -> list[dict[str, object]]:
    sections: list[dict[str, object]] = []
    current: dict[str, object] | None = None
    for line in text.splitlines():
        stripped = line.strip()
        match = HEADING_RE.match(stripped)
        if match:
            if current is not None:
                sections.append(current)
            current = {
                "number": match.group(1),
                "title": strip_trailing_pageno(match.group(2).strip()),
                "lines": [],
            }
            continue
        if current is not None:
            current["lines"].append(line)
    if current is not None:
        sections.append(current)
    return sections


def split_columns(line: str) -> list[str]:
    return [part.strip() for part in re.split(r"\t+|  +", line.strip()) if part.strip()]


def first_non_empty(lines: Iterable[str]) -> str:
    for line in lines:
        stripped = line.strip()
        if stripped:
            return stripped
    return ""


def dedup_keep_order(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def split_subsections(
    lines: list[str],
    markers: tuple[str, ...],
    marker_re: re.Pattern[str],
) -> dict[str, list[str]]:
    subsections: dict[str, list[str]] = {marker: [] for marker in markers}
    current: str | None = None
    for line in lines:
        stripped = line.strip()
        match = marker_re.match(stripped)
        if match:
            current = match.group(1)
            continue
        if current is not None:
            subsections[current].append(line)
    return subsections


def iter_tables(lines: list[str]):
    i = 0
    while i < len(lines):
        stripped = lines[i].strip()
        if not TABLE_TITLE_RE.match(stripped):
            i += 1
            continue
        title = stripped
        body: list[str] = []
        j = i + 1
        while j < len(lines):
            candidate = lines[j].strip()
            if TABLE_TITLE_RE.match(candidate):
                break
            if ANY_TABLE_HEADER_RE.match(candidate):
                break
            if candidate in PODM_MARKERS or candidate in BMC_MARKERS:
                break
            if HEADING_RE.match(candidate):
                break
            if STATUS_CODE_PROSE_RE.match(candidate):
                break
            body.append(lines[j])
            j += 1
        yield title, body
        i = j


def looks_like_type(value: str) -> bool:
    if not value:
        return False
    head = re.split(r"[,，、;；:：\s]", value.strip(), maxsplit=1)[0]
    lowered = head.lower()
    return lowered in TYPE_VALUES or head in TYPE_VALUES or bool(TYPE_SUFFIX_RE.match(head))


def normalize_type(value: str) -> str:
    stripped = value.strip()
    if not stripped:
        return ""
    head = re.split(r"[,，、;；:：\s]", stripped, maxsplit=1)[0]
    if looks_like_type(head):
        return head
    return stripped


def looks_like_name(value: str) -> bool:
    return bool(NAME_LIKE_RE.match(value.strip()))


def classify_podm_table(title: str) -> str:
    lowered = title.lower()
    if "path" in lowered:
        return "path"
    if "header" in lowered:
        return "header"
    if "body" in lowered:
        return "body"
    if "query" in lowered:
        return "query"
    if "response" in lowered or "响应" in title or "返回" in title:
        return "response"
    return "body"


def parse_param_table(body_lines: list[str]) -> list[ParamRecord]:
    header: list[str] | None = None
    records: list[ParamRecord] = []

    for line in body_lines:
        if not line.strip():
            continue
        cols = split_columns(line)
        if not cols:
            continue
        if header is None:
            header = cols
            continue
        if is_table_header(cols):
            header = cols
            continue
        record = parse_param_row(cols, header)
        if record is not None:
            records.append(record)

    return records


def is_table_header(cols: list[str]) -> bool:
    joined = "".join(cols)
    return (
        "参数名称" in joined
        or "字段" in cols[:2]
        or ("参数" in cols[:2] and ("类型" in joined or "说明" in joined))
    )


def parse_param_row(cols: list[str], header: list[str]) -> ParamRecord | None:
    name_index = find_column(header, ("参数名称", "参数名", "参数", "字段"))
    type_index = find_column(header, ("类型", "参数类型", "数据类型"))
    required_index = find_column(header, ("必选", "是否必选", "必选项"))
    description_index = find_column(header, ("参数说明", "说明", "描述"))
    value_index = find_column(header, ("取值", "参数值域", "值域", "默认值"))

    if name_index is None:
        name_index = 0
    if type_index is None:
        type_index = infer_type_index(cols)

    name = safe_get(cols, name_index)
    if not looks_like_name(name):
        return None

    raw_type = safe_get(cols, type_index) if type_index is not None else ""
    return ParamRecord(
        name=name,
        type=normalize_type(raw_type),
        required=safe_get(cols, required_index),
        description=safe_get(cols, description_index),
        value=safe_get(cols, value_index),
    )


def find_column(header: list[str], names: tuple[str, ...]) -> int | None:
    for index, col in enumerate(header):
        normalized = col.replace(" ", "")
        if any(name in normalized for name in names):
            return index
    return None


def infer_type_index(cols: list[str]) -> int | None:
    for index in range(1, min(len(cols), 4)):
        if looks_like_type(cols[index]):
            return index
    return None


def safe_get(cols: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(cols):
        return ""
    return cols[index].strip()


def parse_command_format(lines: list[str]) -> dict[str, object]:
    method = ""
    url = ""

    for line in lines:
        match = METHOD_LINE_RE.match(line.strip())
        if match:
            method = match.group(1).upper()
            break

    for i, line in enumerate(lines):
        match = URL_LINE_RE.match(line.strip())
        if not match:
            continue
        inline = match.group(1).strip()
        if inline:
            url = inline
        else:
            url = first_non_empty(lines[i + 1 :])
        break

    headers: list[str] = []
    body_lines: list[str] = []
    in_headers = False
    in_body = False
    for line in lines:
        stripped = line.strip()
        if re.match(r"^请求头\s*[:：]?\s*$", stripped):
            in_headers = True
            in_body = False
            continue
        if re.match(r"^请求消息体\s*[:：]?\s*$", stripped):
            in_headers = False
            in_body = True
            continue
        if re.match(r"^请求消息体\s*[:：]\s*无\s*$", stripped):
            in_headers = False
            in_body = False
            continue
        if stripped in BMC_MARKERS:
            break
        if in_headers:
            header_match = HEADER_LINE_RE.match(stripped)
            if header_match:
                headers.append(header_match.group(1))
        if in_body:
            body_lines.append(line)

    body_text = "\n".join(body_lines)
    body_keys = [match.group(1) for match in JSON_KEY_RE.finditer(body_text)]

    return {
        "method": method,
        "url": url,
        "headers": dedup_keep_order(headers),
        "body_keys": dedup_keep_order(body_keys),
        "query_keys": extract_query_keys(url),
        "path_keys": extract_path_keys(url),
    }


def extract_query_keys(url: str) -> list[str]:
    if "?" not in url:
        return []
    query = url.split("?", 1)[1]
    names: list[str] = []
    for pair in query.split("&"):
        if not pair:
            continue
        key, sep, value = pair.partition("=")
        if not sep:
            names.append(key)
            continue
        placeholders = query_value_placeholders(value)
        names.extend(placeholders or [key])
    return dedup_keep_order(name for name in names if name)


def query_value_placeholders(value: str) -> list[str]:
    stripped = value.strip()
    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", stripped):
        return [stripped]
    return re.findall(r"=\s*([A-Za-z_][A-Za-z0-9_]*)", stripped)


def extract_path_keys(url: str) -> list[str]:
    if not url:
        return []
    path = re.sub(r"^https?://[^/]+", "", url).split("?", 1)[0]
    braced = BRACED_PLACEHOLDER_RE.findall(path)
    if braced:
        return dedup_keep_order(braced)
    out: list[str] = []
    for segment in path.strip("/").split("/"):
        if not segment or segment.lower() in PATH_KEYWORDS:
            continue
        if segment.startswith("$") or "." in segment:
            continue
        if "_" in segment or segment.endswith(("Id", "ID", "_id")) or segment.islower():
            out.append(segment)
    return dedup_keep_order(out)


def parse_bmc_request_params(
    lines: list[str],
    path_keys: list[str],
    header_keys: list[str],
    body_keys: list[str],
    query_keys: list[str],
) -> dict[str, list[ParamRecord]]:
    params = empty_params()
    rows = parse_param_table(lines)
    key_to_category: dict[str, str] = {}
    for key in path_keys:
        key_to_category[key.lower()] = "path"
    for key in header_keys:
        key_to_category[key.lower()] = "header"
    for key in body_keys:
        key_to_category[key.lower()] = "body"
    for key in query_keys:
        key_to_category[key.lower()] = "query"

    for row in rows:
        if row.name.lower() in IGNORED_PARAM_NAMES:
            continue
        category = key_to_category.get(row.name.lower(), "")
        if not category:
            category = "body"
        params[category].append(row)

    ensure_inferred_params(params, "path", path_keys)
    ensure_inferred_params(params, "header", header_keys)
    ensure_inferred_params(params, "body", body_keys)
    ensure_inferred_params(params, "query", query_keys)
    return params


def ensure_inferred_params(
    params: dict[str, list[ParamRecord]], category: str, names: list[str]
) -> None:
    existing = {record.name.lower() for record in params[category]}
    for name in names:
        if name.lower() not in existing:
            params[category].append(ParamRecord(name=name, inferred=True))


def parse_response_params(lines: list[str]) -> list[ParamRecord]:
    records: list[ParamRecord] = []
    for line in lines:
        cols = split_columns(line)
        if len(cols) < 2:
            continue
        if is_table_header(cols):
            continue
        name, raw_type = cols[0], cols[1]
        if looks_like_name(name) and looks_like_type(raw_type):
            description = cols[2] if len(cols) > 2 else ""
            records.append(
                ParamRecord(
                    name=name,
                    type=normalize_type(raw_type),
                    description=description,
                )
            )
    return records


def empty_params() -> dict[str, list[ParamRecord]]:
    return {category: [] for category in CATEGORIES}


def extract_podm_interface(section: dict[str, object], index: int) -> InterfaceRecord | None:
    lines = list(section["lines"])
    if not any(line.strip() == "URI" for line in lines):
        return None
    subsections = split_subsections(lines, PODM_MARKERS, PODM_MARKER_RE)
    uri = first_non_empty(subsections["URI"])
    if not uri:
        return None
    iface = InterfaceRecord(
        index=index,
        section=str(section["number"]),
        title=str(section["title"]),
        method=first_non_empty(subsections["调用方法"]).upper(),
        uri=uri,
        params=empty_params(),
    )
    for title, body in iter_tables(subsections["请求参数"]):
        category = classify_podm_table(title)
        if category in iface.params:
            iface.params[category].extend(parse_param_table(body))
    for _title, body in iter_tables(subsections["响应参数"]):
        iface.params["response"].extend(parse_param_table(body))
    ensure_inferred_params(iface.params, "path", extract_path_keys(uri))
    ensure_inferred_params(iface.params, "query", extract_query_keys(uri))
    return iface


def extract_bmc_interface(section: dict[str, object], index: int) -> InterfaceRecord | None:
    lines = list(section["lines"])
    if not any(line.strip() == "命令格式" for line in lines):
        return None
    subsections = split_subsections(lines, BMC_MARKERS, BMC_MARKER_RE)
    command = parse_command_format(subsections["命令格式"])
    uri = str(command["url"])
    if not uri:
        return None
    params = parse_bmc_request_params(
        subsections["参数说明"],
        list(command["path_keys"]),
        list(command["headers"]),
        list(command["body_keys"]),
        list(command["query_keys"]),
    )
    params["response"].extend(parse_response_params(subsections["输出说明"]))
    return InterfaceRecord(
        index=index,
        section=str(section["number"]),
        title=str(section["title"]),
        method=str(command["method"]),
        uri=uri,
        params=params,
    )


def extract_interfaces_from_text(
    text: str,
    profile: str,
    section_filter: set[str] | None = None,
) -> list[InterfaceRecord]:
    sections = split_sections(text)
    if profile == "auto":
        profile = detect_profile(sections)
    records: list[InterfaceRecord] = []
    for section in sections:
        section_number = str(section["number"])
        if section_filter is not None and section_number not in section_filter:
            continue
        extractor = extract_bmc_interface if profile == "bmc" else extract_podm_interface
        iface = extractor(section, len(records) + 1)
        if iface is not None:
            records.append(iface)
    return records


def detect_profile(sections: list[dict[str, object]]) -> str:
    bmc_hits = 0
    podm_hits = 0
    for section in sections:
        lines = [line.strip() for line in list(section["lines"])]
        if "命令格式" in lines:
            bmc_hits += 1
        if "URI" in lines and "调用方法" in lines:
            podm_hits += 1
    return "bmc" if bmc_hits > podm_hits else "podm"


def load_common_sections(workbook: Path, side: str, sheet_name: str) -> set[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("读取 xlsx 需要 openpyxl：pip install openpyxl") from exc

    wb = load_workbook(workbook, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Excel 不存在 sheet: {sheet_name}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(value) if value is not None else "" for value in next(rows)]
    except StopIteration as exc:
        raise SystemExit(f"Excel sheet 为空: {sheet_name}") from exc
    column = f"{side}_section"
    try:
        index = header.index(column)
    except ValueError as exc:
        raise SystemExit(f"Excel sheet 缺少列: {column}") from exc
    sections: set[str] = set()
    for row in rows:
        value = row[index] if index < len(row) else None
        if value is not None and str(value).strip():
            sections.add(str(value).strip())
    return sections


def interface_to_dict(iface: InterfaceRecord) -> dict[str, object]:
    data = asdict(iface)
    data["params"] = {
        category: [compact_param(record) for record in iface.params[category]]
        for category in CATEGORIES
    }
    return data


def compact_param(record: ParamRecord) -> dict[str, object]:
    return {"name": record.name, "type": record.type}


def write_output(data: dict[str, object], output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.suffix.lower() == ".json":
        output.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return output
    try:
        import yaml
    except ImportError:
        json_output = output.with_suffix(".json")
        json_output.write_text(
            json.dumps(data, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        return json_output

    with output.open("w", encoding="utf-8") as handle:
        yaml.safe_dump(data, handle, allow_unicode=True, sort_keys=False)
    return output


def resolve_output(input_path: Path, output: Path | None) -> Path:
    if output is not None:
        return output
    return input_path.with_name(f"{input_path.stem}.word.interface-params.yaml")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="从 Redfish Word 接口文档抽取接口、参数名和参数类型"
    )
    parser.add_argument("input", type=Path, help="Word .docx 或展开后的 .txt 文档")
    parser.add_argument(
        "--profile",
        choices=("auto", "podm", "bmc"),
        default="auto",
        help="文档格式；auto 会按 marker 自动判断",
    )
    parser.add_argument("-o", "--output", type=Path, help="输出 .yaml 或 .json 路径")
    parser.add_argument(
        "--match-workbook",
        type=Path,
        help="可选：interface_match_llm_summary.xlsx，只抽取共有接口 sheet 中的 section",
    )
    parser.add_argument(
        "--match-side",
        choices=("auto", "bmc", "podm"),
        default="auto",
        help="匹配表使用哪一侧 section；默认跟 profile 一致",
    )
    parser.add_argument("--match-sheet", default="共有接口", help="匹配表 sheet 名")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    if not args.input.exists():
        raise SystemExit(f"输入文件不存在: {args.input}")

    text = read_source(args.input)
    profile = args.profile
    if profile == "auto":
        profile = detect_profile(split_sections(text))

    section_filter: set[str] | None = None
    if args.match_workbook is not None:
        side = profile if args.match_side == "auto" else args.match_side
        section_filter = load_common_sections(args.match_workbook, side, args.match_sheet)

    interfaces = extract_interfaces_from_text(text, profile, section_filter)
    data = {
        "source": str(args.input),
        "profile": profile,
        "interface_count": len(interfaces),
        "interfaces": [interface_to_dict(iface) for iface in interfaces],
    }
    final_output = write_output(data, resolve_output(args.input, args.output))

    filter_note = ""
    if section_filter is not None:
        filter_note = f"（匹配表 section {len(section_filter)} 条）"
    print(f"已提取 {len(interfaces)} 个 {profile} 接口{filter_note} -> {final_output}")


if __name__ == "__main__":
    main()
